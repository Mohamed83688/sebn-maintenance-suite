import pandas as pd
import os
import openpyxl
import re
import datetime
from datetime import datetime as dt_class

class DataEngine:
    """
    Handles Excel reading and data processing for the Preventive Maintenance System.
    Optimized for SEBN-TN Excel structures.
    """
    
    @staticmethod
    def get_month_map(year=None):
        """Generates a dynamic month-to-week mapping for the specified year using ISO 8601 logic."""
        import datetime
        from collections import defaultdict
        
        if not year:
            year = datetime.datetime.now().year
            
        month_weeks = defaultdict(set)
        # Check every day of the year
        start_date = datetime.date(year, 1, 1)
        end_date = datetime.date(year, 12, 31)
        
        curr = start_date
        while curr <= end_date:
            iso_year, iso_week, iso_day = curr.isocalendar()
            # We only care about weeks in the target year (ISO weeks can span years)
            if iso_year == year:
                # Standard convention: Week belongs to the month containing its Thursday (iso_day 4)
                # or simpler: week belongs to the month where most of its days are.
                # However, for industry dashboards, we usually map a week based on its Thursday.
                thursday = curr + datetime.timedelta(days=(4 - iso_day))
                target_month = thursday.strftime("%B")
                month_weeks[target_month].add(iso_week)
            curr += datetime.timedelta(days=1)
            
        # Convert sets to sorted ranges/lists
        result = {}
        for mname in ["January", "February", "March", "April", "May", "June", 
                      "July", "August", "September", "October", "November", "December"]:
            weeks = sorted(list(month_weeks.get(mname, [])))
            result[mname] = weeks
        return result

    @staticmethod
    def get_month_from_week(week_num, year=None):
        m_map = DataEngine.get_month_map(year)
        for mname, weeks in m_map.items():
            if week_num in weeks:
                return mname
        return "Unknown"

    def __init__(self):
        self.current_df = None
        self.all_sheets = []
        self.project_name = ""

    def _parse_single_excel(self, file_path):
        """
        Helper to parse a single Excel file with explicit header detection and column mapping.
        Guarantees that only valid machines from designated equipment columns are parsed.
        """
        if not os.path.exists(file_path):
            return []

        import re
        match = re.search(r'20\d{2}', os.path.basename(file_path))
        active_year = int(match.group()) if match else datetime.datetime.now().year
        
        ext = os.path.splitext(file_path)[1].lower()
        
        try:
            engine = 'calamine'
            pd.read_excel(file_path, engine=engine, nrows=1)
        except Exception:
            engine = 'openpyxl' if ext in ['.xlsx', '.xlsm'] else 'xlrd'

        tasks = []
        
        try:
            with pd.ExcelFile(file_path, engine=engine) as xl:
                # Blacklist known non-schedule sheets
                blacklist_sheets = [
                    "SOMMAIRE", "HELP", "LEGENDE", "MODE", "ANNEXE",
                    "SIGNATURE", "SIGNATURES", "USER", "USERS", "ACCOUNT",
                    "LOGIN", "TECHNICIEN", "PARAM", "CONFIG", "BACKUP", "TEMPLATE", "FEUIL"
                ]
                sheets_to_process = []
                for s in xl.sheet_names:
                    s_up = s.upper().strip()
                    if any(x in s_up for x in blacklist_sheets):
                        continue
                    if s_up == "TABLE" or s_up.startswith("TABLE "):
                        continue
                    sheets_to_process.append(s)

                for sheet_name in sheets_to_process:
                    self.current_sheet_name = sheet_name
                    sheet_data = pd.read_excel(xl, sheet_name=sheet_name, header=None)
                    
                    if sheet_data.empty or len(sheet_data) < 2:
                        continue

                    # 1. Detect exact Header Row containing week columns (KW 1, S1, W1...)
                    hdr_row_idx = None
                    week_start_col = None
                    
                    for r_i in range(min(20, len(sheet_data))):
                        row_vals = [str(x).upper().strip() for x in sheet_data.iloc[r_i]]
                        for c_i, val in enumerate(row_vals):
                            if val in ["S1", "S01", "W1", "W01", "KW 1", "KW1", "KW01", "KW 01", "KW.1", "WEEK 1"]:
                                hdr_row_idx = r_i
                                week_start_col = c_i
                                break
                            elif val.startswith("KW ") or val.startswith("S ") or val.startswith("W "):
                                suffix = val.split(" ")[-1]
                                if suffix.isdigit() and int(suffix) == 1:
                                    hdr_row_idx = r_i
                                    week_start_col = c_i
                                    break
                        if hdr_row_idx is not None:
                            break

                    # If this sheet doesn't have a week header row, it's not a schedule sheet
                    if hdr_row_idx is None or week_start_col is None:
                        continue

                    # 2. Identify exact column positions from the header row
                    hdr_row = [str(sheet_data.iloc[hdr_row_idx, c] or '').strip() for c in range(week_start_col)]
                    col_machine = None
                    col_name = None
                    col_group = None
                    col_carte = None

                    for c_i, h_val in enumerate(hdr_row):
                        h_up = h_val.upper()
                        if any(k in h_up for k in ['SEMAINE', 'ID MACHINE', 'MACHINE', 'EQUIPMENT', 'EQUIPEMENT']):
                            col_machine = c_i
                        elif h_up in ['N°', 'NO', 'DESIGNATION', 'NOM MACHINE']:
                            col_name = c_i
                        elif any(k in h_up for k in ['ZONE', 'GROUPE', 'GROUP']):
                            col_group = c_i
                        elif any(k in h_up for k in ['CARTE', 'MATRICULE']):
                            col_carte = c_i

                    # Fallback if machine column header wasn't labeled: column immediately before week_start_col
                    if col_machine is None:
                        col_machine = week_start_col - 1

                    # 3. Parse data rows below the header row
                    for r_i in range(hdr_row_idx + 1, len(sheet_data)):
                        row = sheet_data.iloc[r_i]
                        if len(row) <= week_start_col:
                            continue

                        # Extract machine strictly from col_machine
                        raw_machine = str(row[col_machine]).strip() if col_machine < len(row) and pd.notna(row[col_machine]) else ""
                        
                        # Validate machine ID: MUST NOT be empty, nan, or non-machine keyword
                        if not raw_machine or raw_machine.lower() in ['nan', 'none', '']:
                            continue

                        raw_machine_up = raw_machine.upper()
                        noise_keywords = [
                            'ZONE', 'N° CARTE', 'SEMAINE', 'TOTAL', 'ROLE', 'ADMIN', 'ADMINISTRATEUR',
                            'SIGNATURES', 'SIGNATURE', 'TECHNICIAN', 'TECHNICIEN', 'USER', 'USERS',
                            'ACCOUNT', 'LOGIN', 'MATRICULE', 'DATE', 'SHIFT', 'EQUIPE', 'PAGE',
                            'RESP', 'REV', 'ANNEXE', 'SOMMAIRE', 'VALIDÉ', 'VALIDE', 'VISA'
                        ]
                        if raw_machine_up in noise_keywords or raw_machine_up.startswith('PPE-VA') or raw_machine_up.startswith('ANNEXE'):
                            continue

                        # Extract group, machine name, and carte
                        zone_val = str(row[col_group]).strip() if col_group is not None and col_group < len(row) and pd.notna(row[col_group]) else ""
                        group_name = zone_val if zone_val and zone_val.lower() != 'nan' else sheet_name.strip()
                        
                        m_name_val = str(row[col_name]).strip() if col_name is not None and col_name < len(row) and pd.notna(row[col_name]) else ""
                        machine_name = m_name_val if m_name_val and m_name_val.lower() != 'nan' else raw_machine

                        carte_val = str(row[col_carte]).strip() if col_carte is not None and col_carte < len(row) and pd.notna(row[col_carte]) else ""
                        carte = carte_val if carte_val and carte_val.lower() != 'nan' else ""

                        # Extract tasks from week columns
                        start_c = int(week_start_col)
                        limit_c = min(start_c + 54, int(len(row)))
                        for col_idx in range(start_c, limit_c):
                            val = row[col_idx]
                            cell_val = str(val).strip().upper() if pd.notna(val) else ""
                            
                            if any(x in cell_val for x in ['H', 'M', 'DONE', 'X', 'OK']):
                                week_num = int(col_idx - start_c + 1)
                                month_name = self.get_month_from_week(week_num, active_year)
                                
                                task_type = "Weekly" if 'H' in cell_val else "Monthly"
                                is_done = any(x in cell_val for x in ["DONE", "X", "OK"])
                                
                                tasks.append({
                                    "Equipment": raw_machine,
                                    "Machine_Name": machine_name,
                                    "Zone": group_name,
                                    "Group": group_name,
                                    "Sheet": sheet_name,
                                    "Carte": carte,
                                    "Matricule": carte,
                                    "Week": f"S{week_num}",
                                    "Month": month_name,
                                    "Type": task_type,
                                    "Status": "COMPLÉTÉ" if is_done else "INCOMPLÉTÉ",
                                    "Raw_Index": (r_i, col_idx),
                                    "Source_File": file_path
                                })
        except Exception as e:
            print(f"Error parsing Excel file {file_path}: {e}")

        return tasks

    def load_excel(self, file_path):
        """
        Reads one or multiple Excel files and extracts maintenance tasks.
        Accepts a single file path (str) or a list/tuple of file paths.
        Guarantees deduplication so identical tasks are not multiplied.
        """
        if isinstance(file_path, (list, tuple)):
            valid_paths = [p for p in file_path if p and os.path.exists(p)]
            if not valid_paths:
                self.current_df = pd.DataFrame()
                return self.current_df
            
            all_tasks = []
            seen_keys = set()
            for fp in valid_paths:
                tasks_fp = self._parse_single_excel(fp)
                for t in tasks_fp:
                    key = (
                        str(t.get('Equipment', '')).strip().upper(),
                        str(t.get('Sheet', '')).strip().upper(),
                        str(t.get('Week', '')).strip().upper()
                    )
                    if key not in seen_keys:
                        seen_keys.add(key)
                        all_tasks.append(t)
                
            self.source_file = valid_paths[0]
            self.source_files = valid_paths
            self.current_df = pd.DataFrame(all_tasks)
            return self.current_df
        else:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            raw_tasks = self._parse_single_excel(file_path)
            all_tasks = []
            seen_keys = set()
            for t in raw_tasks:
                key = (
                    str(t.get('Equipment', '')).strip().upper(),
                    str(t.get('Sheet', '')).strip().upper(),
                    str(t.get('Week', '')).strip().upper()
                )
                if key not in seen_keys:
                    seen_keys.add(key)
                    all_tasks.append(t)
            self.source_file = file_path
            self.source_files = [file_path]
            self.current_df = pd.DataFrame(all_tasks)
            return self.current_df

    def get_source_file_for_task(self, sheet_name, raw_idx):
        """Finds the source Excel file path for a specific task identified by sheet and raw_idx."""
        if isinstance(raw_idx, str):
            import ast
            try: raw_idx = ast.literal_eval(raw_idx)
            except: pass
        if self.current_df is not None and not self.current_df.empty and 'Source_File' in self.current_df.columns:
            match = self.current_df[
                (self.current_df['Sheet'] == sheet_name) & 
                (self.current_df['Raw_Index'].apply(lambda x: tuple(x) == tuple(raw_idx)))
            ]
            if not match.empty:
                return match.iloc[0]['Source_File']
        return getattr(self, 'source_file', None)

    def save_task_done(self, sheet_name, raw_idx, file_path=None):
        """Writes 'DONE' status (X) directly back to the physical Excel file."""
        target_file = file_path or self.get_source_file_for_task(sheet_name, raw_idx)
        if not target_file or not os.path.exists(target_file):
            return False
        try:
            ext = os.path.splitext(target_file)[1].lower()
            use_vba = ext in ['.xlsm', '.xltm']
            wb = openpyxl.load_workbook(target_file, keep_vba=use_vba)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                row, col = raw_idx[0] + 1, raw_idx[1] + 1
                ws.cell(row=row, column=col).value = "X"
                wb.save(target_file)
                return True
        except Exception as e:
            print(f"Excel write error for {target_file}: {e}")
        return False

    def complete_task(self, sheet_name, raw_idx, task_type=None, tech_name=None, shift=None):
        if isinstance(raw_idx, str):
            import ast
            try: raw_idx = ast.literal_eval(raw_idx)
            except: pass
        if isinstance(raw_idx, (tuple, list)) and len(raw_idx) == 2:
            target_file = self.get_source_file_for_task(sheet_name, raw_idx)
            ok = self.save_task_done(sheet_name, raw_idx, file_path=target_file)
            if ok and self.current_df is not None and not self.current_df.empty:
                mask = (self.current_df['Sheet'] == sheet_name) & (self.current_df['Raw_Index'].apply(lambda x: tuple(x) == tuple(raw_idx)))
                self.current_df.loc[mask, 'Status'] = 'COMPLÉTÉ'
            return ok
        return False

    def get_stats(self):
        if self.current_df is None or self.current_df.empty:
            return {
                "total": 0, "done": 0, "pending": 0, "rate": 0,
                "monthly_data": [],
                "type_data": {"Monthly": 0, "Weekly": 0, "Quarterly": 0, "Annual": 0},
                "weekly_data": [],
                "by_sheet": {},
                "by_week": {},
                "recent_tasks": []
            }
        
        df = self.current_df
        total = len(df)
        
        done_mask = df['Status'].astype(str).str.upper().str.strip() == 'COMPLÉTÉ'
        done = int(done_mask.sum())
        pending = total - done
        rate = round((done / total) * 100) if total > 0 else 0
        
        monthly_data = []
        if 'Month' in df.columns:
            months = df['Month'].dropna().unique()
            months = sorted([m for m in months if str(m).strip() != ''])
            for m in months[-12:]:
                m_df = df[df['Month'] == m]
                m_total = len(m_df)
                m_done = int((m_df['Status'].astype(str).str.upper().str.strip() == 'COMPLÉTÉ').sum())
                m_rate = round((m_done / m_total) * 100) if m_total > 0 else 0
                monthly_data.append({"month": str(m), "total": m_total, "done": m_done, "rate": m_rate})
                
        type_data = {"Monthly": 0, "Weekly": 0, "Quarterly": 0, "Annual": 0}
        if 'Type' in df.columns:
            counts = df['Type'].value_counts().to_dict()
            for k, v in counts.items():
                k_str = str(k).strip()
                if k_str in type_data:
                    type_data[k_str] += int(v)
                else:
                    k_lower = k_str.lower()
                    if "month" in k_lower or "mensuel" in k_lower: type_data["Monthly"] += int(v)
                    elif "week" in k_lower or "hebdo" in k_lower: type_data["Weekly"] += int(v)
                    elif "quart" in k_lower or "trim" in k_lower: type_data["Quarterly"] += int(v)
                    elif "ann" in k_lower: type_data["Annual"] += int(v)

        # ----- by_sheet breakdown -----
        by_sheet = {}
        if 'Sheet' in df.columns:
            for s in df['Sheet'].dropna().unique():
                s_str = str(s).strip()
                if not s_str:
                    continue
                s_df = df[df['Sheet'] == s]
                s_total = len(s_df)
                s_done = int((s_df['Status'].astype(str).str.upper().str.strip() == 'COMPLÉTÉ').sum())
                by_sheet[s_str] = {"total": s_total, "done": s_done, "rate": round((s_done / s_total) * 100) if s_total > 0 else 0}

        # ----- weekly_data list (for charts that expect list format) -----
        weekly_data = []
        if 'Week' in df.columns:
            weeks_sorted = sorted([str(w).strip() for w in df['Week'].dropna().unique() if str(w).strip()])
            for w_str in weeks_sorted[-16:]:
                w_df = df[df['Week'].astype(str).str.strip() == w_str]
                w_total = len(w_df)
                w_done = int((w_df['Status'].astype(str).str.upper().str.strip() == 'COMPLÉTÉ').sum())
                weekly_data.append({"week": w_str, "total": w_total, "done": w_done})

        # ----- by_week keyed dict -----
        by_week = {}
        if 'Week' in df.columns:
            for w in df['Week'].dropna().unique():
                w_str = str(w).strip()
                if not w_str:
                    continue
                w_df = df[df['Week'] == w]
                w_total = len(w_df)
                w_done = int((w_df['Status'].astype(str).str.upper().str.strip() == 'COMPLÉTÉ').sum())
                by_week[w_str] = {"total": w_total, "done": w_done}

        # ----- recent_tasks (last 20 rows) -----
        recent_tasks = []
        sample = df.tail(20)
        for _, r in sample.iterrows():
            recent_tasks.append({
                "equipment": str(r.get('Equipment', '')),
                "type": str(r.get('Type', '')),
                "week": str(r.get('Week', '')),
                "sheet": str(r.get('Sheet', '')),
                "status": str(r.get('Status', '')).upper().strip()
            })
        
        return {
            "total": total,
            "done": done,
            "pending": pending,
            "rate": rate,
            "monthly_data": monthly_data,
            "type_data": type_data,
            "weekly_data": weekly_data,
            "by_sheet": by_sheet,
            "by_week": by_week,
            "recent_tasks": recent_tasks
        }

    def get_filtered_pma_analytics(self, month=None, week=None, sheet=None, machine=None, status=None, task_type=None) -> dict:
        """Fetch comprehensive filtered metrics, charts, and records for PMA analytics."""
        if self.current_df is None or self.current_df.empty:
            return {
                "kpis": {"total": 0, "done": 0, "pending": 0, "overdue": 0, "rate": 0, "overdue_rate": 0},
                "timeline": [],
                "by_month": [],
                "by_machine": [],
                "by_group": [],
                "schedule_adherence": {"done": 0, "overdue": 0, "pending": 0},
                "tasks": [],
                "filters": {"months": [], "weeks": [], "groups": [], "machines": [], "types": []}
            }

        df = self.current_df.copy()

        # Sanitize machines: NEVER treat admin, role, signatures, nan as machines
        bad_equip_mask = df['Equipment'].astype(str).str.upper().str.strip().isin(['ADMIN', 'ROLE', 'SIGNATURES', 'SIGNATURE', 'TOTAL', 'NAN', 'NONE', ''])
        df = df[~bad_equip_mask].copy()

        # Available options before filtering (for dropdowns)
        def _get_w_num(val):
            m = re.search(r'\d+', str(val))
            return int(m.group(0)) if m else 0

        month_order = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december',
                       'janvier', 'février', 'mars', 'avril', 'mai', 'juin', 'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre']
        
        raw_months = [str(m).strip() for m in df['Month'].dropna().unique() if str(m).strip()]
        available_months = sorted(list(set(raw_months)), key=lambda m: month_order.index(m.lower()) if m.lower() in month_order else 99)
        
        raw_weeks = [str(w).strip() for w in df['Week'].dropna().unique() if str(w).strip()]
        available_weeks = sorted(list(set(raw_weeks)), key=_get_w_num)

        group_col = 'Group' if 'Group' in df.columns else 'Sheet'
        raw_groups = [str(g).strip() for g in df[group_col].dropna().unique() if str(g).strip() and str(g).strip().upper() not in ['SIGNATURES', 'USERS', 'CONFIG', 'ADMIN']]
        available_groups = sorted(list(set(raw_groups)))

        raw_machines = [str(m).strip() for m in df['Equipment'].dropna().unique() if str(m).strip()]
        available_machines = sorted(list(set(raw_machines)))

        raw_types = [str(t).strip() for t in df['Type'].dropna().unique() if str(t).strip()]
        available_types = sorted(list(set(raw_types)))

        # Apply filters
        if month and str(month).strip() not in ('All', '', 'Tous'):
            df = df[df['Month'].astype(str).str.strip().str.lower() == str(month).strip().lower()]
        if week and str(week).strip() not in ('All', '', 'Tous'):
            df = df[df['Week'].astype(str).str.strip().str.upper() == str(week).strip().upper()]
        if sheet and str(sheet).strip() not in ('All', '', 'Tous'):
            df = df[(df['Sheet'].astype(str).str.strip().str.lower() == str(sheet).strip().lower()) |
                    (df['Group'].astype(str).str.strip().str.lower() == str(sheet).strip().lower())]
        if machine and str(machine).strip() not in ('All', '', 'Tous'):
            df = df[df['Equipment'].astype(str).str.strip().str.lower() == str(machine).strip().lower()]
        if status and str(status).strip() not in ('All', '', 'Tous'):
            stat_up = str(status).strip().upper()
            if stat_up in ('COMPLÉTÉ', 'COMPLETE', 'DONE', 'COMPLÉTÉE', 'TERMINE', 'TERMINÉE'):
                df = df[df['Status'].astype(str).str.upper().str.strip() == 'COMPLÉTÉ']
            elif stat_up in ('EN ATTENTE', 'PENDING', 'INCOMPLÉTÉ', 'ATTENTE'):
                df = df[df['Status'].astype(str).str.upper().str.strip() != 'COMPLÉTÉ']
        if task_type and str(task_type).strip() not in ('All', '', 'Tous'):
            df = df[df['Type'].astype(str).str.strip().str.lower() == str(task_type).strip().lower()]

        current_week_num = datetime.datetime.now().isocalendar()[1]

        df['Week_Num'] = df['Week'].apply(_get_w_num)
        df['Is_Done'] = df['Status'].astype(str).str.upper().str.strip() == 'COMPLÉTÉ'
        df['Is_Overdue'] = (~df['Is_Done']) & (df['Week_Num'] < current_week_num)
        df['Is_Pending'] = (~df['Is_Done']) & (df['Week_Num'] >= current_week_num)

        total = len(df)
        done = int(df['Is_Done'].sum())
        overdue = int(df['Is_Overdue'].sum())
        pending = int(df['Is_Pending'].sum())
        rate = round((done / total) * 100, 1) if total > 0 else 0
        overdue_rate = round((overdue / total) * 100, 1) if total > 0 else 0

        # Timeline by week
        timeline = []
        weeks_in_df = sorted([str(w).strip() for w in df['Week'].dropna().unique() if str(w).strip()],
                             key=_get_w_num)
        for w in weeks_in_df:
            w_df = df[df['Week'].astype(str).str.strip() == w]
            timeline.append({
                "week": w,
                "total": len(w_df),
                "done": int(w_df['Is_Done'].sum()),
                "overdue": int(w_df['Is_Overdue'].sum()),
                "pending": int(w_df['Is_Pending'].sum()),
            })

        # By month
        by_month = []
        months_in_df = df['Month'].dropna().unique().tolist()
        months_in_df = sorted(months_in_df, key=lambda m: month_order.index(str(m).strip().lower()) if str(m).strip().lower() in month_order else 99)
        for m in months_in_df:
            m_df = df[df['Month'] == m]
            m_tot = len(m_df)
            m_dn = int(m_df['Is_Done'].sum())
            by_month.append({
                "month": str(m),
                "total": m_tot,
                "done": m_dn,
                "overdue": int(m_df['Is_Overdue'].sum()),
                "rate": round((m_dn / m_tot) * 100, 1) if m_tot > 0 else 0
            })

        # By machine (Top 12 machines)
        by_machine = []
        machine_counts = df['Equipment'].value_counts()
        for equip, tot in machine_counts.head(12).items():
            equip_str = str(equip).strip()
            eq_df = df[df['Equipment'] == equip]
            eq_dn = int(eq_df['Is_Done'].sum())
            eq_name = eq_df['Machine_Name'].dropna().iloc[0] if not eq_df['Machine_Name'].dropna().empty else equip_str
            by_machine.append({
                "equipment": equip_str,
                "machine_name": str(eq_name),
                "total": int(tot),
                "done": eq_dn,
                "pending": int(eq_df['Is_Pending'].sum()),
                "overdue": int(eq_df['Is_Overdue'].sum()),
                "rate": round((eq_dn / int(tot)) * 100, 1) if tot > 0 else 0
            })

        # By group / sheet
        by_group = []
        for grp in df[group_col].dropna().unique():
            grp_str = str(grp).strip()
            if not grp_str or grp_str.upper() in ['SIGNATURES', 'USERS', 'CONFIG', 'ADMIN']:
                continue
            g_df = df[df[group_col] == grp]
            g_tot = len(g_df)
            g_dn = int(g_df['Is_Done'].sum())
            by_group.append({
                "group": grp_str,
                "total": g_tot,
                "done": g_dn,
                "overdue": int(g_df['Is_Overdue'].sum()),
                "pending": int(g_df['Is_Pending'].sum()),
                "rate": round((g_dn / g_tot) * 100, 1) if g_tot > 0 else 0
            })
        by_group = sorted(by_group, key=lambda x: x['total'], reverse=True)

        # Detailed tasks list (for analytical table)
        tasks_list = []
        for _, r in df.head(300).iterrows():
            tasks_list.append({
                "equipment": str(r.get('Equipment', '')),
                "machine_name": str(r.get('Machine_Name', '')),
                "zone": str(r.get('Zone', r.get('Group', ''))),
                "carte": str(r.get('Carte', '')),
                "type": str(r.get('Type', '')),
                "week": str(r.get('Week', '')),
                "month": str(r.get('Month', '')),
                "status": "COMPLÉTÉ" if r.get('Is_Done') else ("EN RETARD" if r.get('Is_Overdue') else "EN ATTENTE"),
                "is_done": bool(r.get('Is_Done')),
                "is_overdue": bool(r.get('Is_Overdue'))
            })

        return {
            "kpis": {
                "total": total,
                "done": done,
                "overdue": overdue,
                "pending": pending,
                "rate": rate,
                "overdue_rate": overdue_rate
            },
            "timeline": timeline,
            "by_month": by_month,
            "by_machine": by_machine,
            "by_group": by_group,
            "schedule_adherence": {
                "done": done,
                "overdue": overdue,
                "pending": pending
            },
            "tasks": tasks_list,
            "filters": {
                "months": available_months,
                "weeks": available_weeks,
                "groups": available_groups,
                "machines": available_machines,
                "types": available_types
            }
        }

    def reset_task_status(self, sheet_name, raw_idx, task_type="Monthly", file_path=None):
        """Reverts 'DONE' status back to its original marker."""
        target_file = file_path or self.get_source_file_for_task(sheet_name, raw_idx)
        if not target_file or not os.path.exists(target_file):
            return False
        try:
            ext = os.path.splitext(target_file)[1].lower()
            use_vba = ext in ['.xlsm', '.xltm']
            wb = openpyxl.load_workbook(target_file, keep_vba=use_vba)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                row, col = raw_idx[0] + 1, raw_idx[1] + 1
                # Restore original marker (H for Weekly, M for Monthly)
                if task_type == "Weekly":
                    val = "H"
                else:
                    val = "M"
                ws.cell(row=row, column=col).value = val
                wb.save(target_file)
                return True
        except Exception as e:
            print(f"Excel reset error: {e}")
        return False

    def _identify_machine(self, row, week_start_col):
        """Helper to find machine name in the columns before week starting."""
        search_range = range(0, week_start_col)
        # Strategy A: Check for "TS" patterns
        for col_idx in search_range:
            val = str(row[col_idx]).strip()
            if "TS" in val.upper() and any(char.isdigit() for char in val):
                return val
        
        # Strategy B: First long string that looks like an ID
        for col_idx in reversed(search_range):
            val = str(row[col_idx]).strip()
            # Relaxed identification: any non-empty string >= 3 chars that isn't blacklisted
            if val and val.lower() != 'nan' and len(val) >= 3:
                # 1. Skip if it's just the sheet name (Group Name)
                if hasattr(self, 'current_sheet_name') and val.upper() == self.current_sheet_name.upper():
                    continue

                # 2. Skip Drawing Numbers (Pattern: typically long, lots of digits and dashes)
                # Example drawing: 12345-6789-00 or 12345678 (8+ digits)
                digit_count = sum(c.isdigit() for c in val)
                
                # Broadened case-insensitive check for common non-machine noise
                noise_keywords = ["DRAWING", "DESSIN", "DWG", "DRW", "NUMÉRO", "NUMBER"]
                val_up = val.upper()
                
                # If it looks like a drawing or has 8+ digits, skip it
                if any(nk in val_up for nk in noise_keywords) or digit_count >= 8:
                    continue

                # 3. Blacklist for machine IDs (Filter out headers and summary text)
                machine_blacklist = [
                    "ZONE", "ANNÉE", "CARTE", "RESP", "PPE-VA", "ANNEXE", "REV",
                    "SHEET", "SOMMAIRE", "TOTAL", "PAGE", "PROJET", "SECTION",
                    "SEMAINE", "DOMAINE", "DOMAIN", "N°", "NO.", "DESIGNATION",
                    "SOMAIN", "GROUP", "MACHINE", "EQUIPEMENT", "EQUIPMENT", "TABLE",
                    "KSK", "PLAN", "SCHÉMA", "SCHEMA", "REF", "REFERENCE", "TRAÇABILITÉ",
                    # Metadata / signature rows — must NOT be treated as machine IDs
                    "ROLE", "TECHNICIAN", "TECHNICIEN", "SIGNATURES", "SIGNATURE",
                    "VALIDÉ", "VALIDE", "VISA", "APPROUVER", "APPROUVÉ",
                    "NOM", "PRENOM", "MATRICULE", "DATE", "SHIFT", "EQUIPE",
                ]
                if not any(x in val_up for x in machine_blacklist):
                    return val
        return None

    def _check_if_yellow(self, ws, row_idx_0, col_idx_0):
        """Checks if a cell is yellow using openpyxl."""
        try:
            # openpyxl is 1-indexed
            row_idx, col_idx = row_idx_0 + 1, col_idx_0 + 1
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Merged cell handling
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
            
            if cell.fill and hasattr(cell.fill, 'start_color'):
                rgb = str(cell.fill.start_color.rgb).upper()
                # Yellow: FFFF00, Gold: FFD700, Light Yellow: FFF2CC
                if any(y in rgb for y in ["FFFF00", "FFD7", "FFF2", "E26B0A"]): # Added orange-ish check
                    return True
        except:
            pass
        return False
    
    def add_excel_signature(self, file_path, signer_name, role="Technician", matricule=""):
        """
        Add digital signature to Excel file.
        Creates a 'Signatures' sheet with validation history.
        
        Args:
            file_path: Path to Excel file
            signer_name: Name of person signing
            role: Role (Technician/Administrator)
            matricule: Technician matricule/badge number
        
        Returns:
            bool: Success status
        """
        try:
            # DETECTION: Enable VBA preservation for .xlsm files
            ext = os.path.splitext(file_path)[1].lower()
            use_vba = ext in ['.xlsm', '.xltm']
            
            wb = openpyxl.load_workbook(file_path, keep_vba=use_vba)
            
            # Create or get Signatures sheet
            if 'Signatures' not in wb.sheetnames:
                sig_sheet = wb.create_sheet('Signatures')
                
                # Header row with styling — now includes Matricule column
                headers = ['Date', 'Time', 'Role', 'Name', 'Matricule', 'Action']
                for col_idx, header in enumerate(headers, 1):
                    cell = sig_sheet.cell(row=1, column=col_idx, value=header)
                    cell.font = openpyxl.styles.Font(bold=True, size=12)
                    cell.fill = openpyxl.styles.PatternFill(start_color="F0883E", end_color="F0883E", fill_type="solid")
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                
                # Set column widths
                sig_sheet.column_dimensions['A'].width = 15
                sig_sheet.column_dimensions['B'].width = 12
                sig_sheet.column_dimensions['C'].width = 15
                sig_sheet.column_dimensions['D'].width = 25
                sig_sheet.column_dimensions['E'].width = 18
                sig_sheet.column_dimensions['F'].width = 30
            else:
                sig_sheet = wb['Signatures']
                # If existing sheet has old 5-column format, add Matricule header if missing
                existing_headers = [sig_sheet.cell(row=1, column=c).value for c in range(1, 7)]
                if 'Matricule' not in existing_headers:
                    # Shift existing column E data to column F and add Matricule header at E
                    # Only update the header; data rows keep their structure to avoid corruption
                    if sig_sheet.cell(row=1, column=5).value and sig_sheet.cell(row=1, column=5).value != 'Matricule':
                        sig_sheet.insert_cols(5)
                        sig_sheet.cell(row=1, column=5).value = 'Matricule'
                        sig_sheet.cell(row=1, column=5).font = openpyxl.styles.Font(bold=True, size=12)
                        sig_sheet.cell(row=1, column=5).fill = openpyxl.styles.PatternFill(start_color="F0883E", end_color="F0883E", fill_type="solid")
                        sig_sheet.cell(row=1, column=5).alignment = openpyxl.styles.Alignment(horizontal='center')
                        sig_sheet.column_dimensions['E'].width = 18
            
            # Add new signature entry
            next_row = sig_sheet.max_row + 1
            now = datetime.now()
            
            sig_sheet.cell(row=next_row, column=1, value=now.strftime('%Y-%m-%d'))
            sig_sheet.cell(row=next_row, column=2, value=now.strftime('%H:%M:%S'))
            sig_sheet.cell(row=next_row, column=3, value=role)
            sig_sheet.cell(row=next_row, column=4, value=signer_name)
            sig_sheet.cell(row=next_row, column=5, value=matricule or "")  # Matricule column
            sig_sheet.cell(row=next_row, column=6, value=f"Validated by {role}")
            
            # Style the new row
            for col_idx in range(1, 7):
                cell = sig_sheet.cell(row=next_row, column=col_idx)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                if role == "Administrator":
                    cell.fill = openpyxl.styles.PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
            
            wb.save(file_path)
            return True
            
        except Exception as e:
            print(f"Error adding Excel signature: {e}")
            return False
    
    def get_excel_signatures(self, file_path):
        """
        Retrieve all signatures from an Excel file.
        
        Args:
            file_path: Path to Excel file
        
        Returns:
            list: List of signature dictionaries
        """
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            
            if 'Signatures' not in wb.sheetnames:
                return []
            
            sig_sheet = wb['Signatures']
            signatures = []
            
            for row_idx in range(2, sig_sheet.max_row + 1):
                col5_val = sig_sheet.cell(row=row_idx, column=5).value
                col6_val = sig_sheet.cell(row=row_idx, column=6).value if sig_sheet.max_column >= 6 else None
                
                # Handle legacy 5-column vs new 6-column signatures
                if col6_val is not None:
                    mat_val = col5_val
                    act_val = col6_val
                else:
                    mat_val = ""
                    act_val = col5_val

                sig = {
                    'date': sig_sheet.cell(row=row_idx, column=1).value,
                    'time': sig_sheet.cell(row=row_idx, column=2).value,
                    'role': sig_sheet.cell(row=row_idx, column=3).value,
                    'name': sig_sheet.cell(row=row_idx, column=4).value,
                    'matricule': mat_val,
                    'action': act_val
                }
                signatures.append(sig)
            
            wb.close()
            return signatures
            
        except Exception as e:
            print(f"Error reading Excel signatures: {e}")
            return []
