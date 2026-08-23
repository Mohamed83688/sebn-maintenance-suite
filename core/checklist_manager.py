"""
checklist_manager.py
====================
SQLite-backed Checklist Definition, Versioning, and History Management.
Controls the lifecycle of Excel-imported maintenance checklists.

Architecture:
- Excel file -> Validation -> Importer -> Database (ChecklistDefinition -> ChecklistVersion -> ChecklistItem)
- Active configuration -> Only 1 ACTIVE version per definition.
- Deactivation (Remove) -> Status set to INACTIVE (Soft-delete). Historical records are NEVER deleted.
- Executions -> ChecklistExecution records linked to specific Version ID and preserved permanently.
"""

import os
import re
import json
import sqlite3
import hashlib
import datetime
import shutil
import logging
from typing import List, Dict, Tuple, Optional, Any

logger = logging.getLogger("sebn-checklist")

class ChecklistManager:
    def __init__(self, db_path: str, storage_dir: str):
        self.db_path = db_path
        self.storage_dir = storage_dir
        self.archive_dir = os.path.join(storage_dir, "archive")
        
        os.makedirs(self.storage_dir, exist_ok=True)
        os.makedirs(self.archive_dir, exist_ok=True)
        
        self._init_db()
        self._seed_existing_templates()

    def _get_conn(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        return conn

    def _init_db(self):
        """Create database tables if they do not exist."""
        with self._get_conn() as conn:
            cur = conn.cursor()
            
            # 1. Checklist Definitions (Grouping e.g. "PPE Mensuelle Presse TS1300")
            cur.execute("""
            CREATE TABLE IF NOT EXISTS checklist_definitions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                checklist_code TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                target_type TEXT NOT NULL DEFAULT 'PPE',
                equipment_pattern TEXT NOT NULL DEFAULT 'ALL',
                description TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            """)

            # 2. Checklist Versions (V1, V2, V3... with ACTIVE/INACTIVE status)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS checklist_versions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                definition_id INTEGER NOT NULL,
                version_number INTEGER NOT NULL,
                status TEXT NOT NULL DEFAULT 'ACTIVE', -- 'ACTIVE' or 'INACTIVE'
                original_filename TEXT NOT NULL,
                stored_filename TEXT NOT NULL,
                file_sha256 TEXT,
                item_count INTEGER NOT NULL DEFAULT 0,
                imported_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                imported_by TEXT DEFAULT 'admin',
                change_summary TEXT,
                FOREIGN KEY (definition_id) REFERENCES checklist_definitions(id) ON DELETE RESTRICT,
                UNIQUE (definition_id, version_number)
            )
            """)

            # 3. Checklist Items (Inspection points per version)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS checklist_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                version_id INTEGER NOT NULL,
                item_number INTEGER NOT NULL,
                section_header TEXT,
                description TEXT NOT NULL,
                method TEXT,
                control_type TEXT DEFAULT 'OK_NOK_NA',
                icon TEXT DEFAULT 'fa-clipboard-check',
                row_index INTEGER,
                FOREIGN KEY (version_id) REFERENCES checklist_versions(id) ON DELETE RESTRICT
            )
            """)

            # 4. Checklist Executions (Historical completed inspections by technicians)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS checklist_executions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                version_id INTEGER,
                equipment TEXT NOT NULL,
                task_type TEXT NOT NULL,
                sheet TEXT,
                week TEXT,
                month TEXT,
                technician_name TEXT NOT NULL,
                technician_matricule TEXT,
                shift TEXT,
                executed_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                status TEXT DEFAULT 'COMPLETED',
                answers_json TEXT NOT NULL,
                filled_excel_path TEXT,
                FOREIGN KEY (version_id) REFERENCES checklist_versions(id) ON DELETE SET NULL
            )
            """)
            conn.commit()

    def _seed_existing_templates(self):
        """Auto-import any existing .xlsx / .xls files in storage_dir on startup if not yet registered."""
        try:
            with self._get_conn() as conn:
                cur = conn.cursor()
                cur.execute("SELECT COUNT(*) as count FROM checklist_versions")
                row = cur.fetchone()
                if row and row["count"] > 0:
                    return # Already seeded
            
            # Discover existing templates
            if not os.path.isdir(self.storage_dir):
                return
            
            candidates = [
                f for f in os.listdir(self.storage_dir)
                if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith(('~$', 'archive'))
            ]
            for f in sorted(candidates):
                fpath = os.path.join(self.storage_dir, f)
                if os.path.isfile(fpath):
                    self.validate_and_import_excel(
                        file_path=fpath,
                        original_filename=f,
                        imported_by='system_seed'
                    )
        except Exception as e:
            logger.warning(f"Checklist initial seed notice: {e}")

    # =========================================================================
    # Validation & Import
    # =========================================================================

    def validate_excel_file(self, file_path: str) -> Tuple[bool, str, List[Dict[str, Any]], Dict[str, Any]]:
        """
        Validates the Excel checklist structure without modifying database.
        Returns: (is_valid, message, parsed_items, metadata)
        """
        try:
            from web_portal.utils.checklist_parser import parse_tasks, _get_best_sheet
            import openpyxl

            if not os.path.exists(file_path):
                return False, "Fichier introuvable.", [], {}

            # Attempt to open workbook
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
            except Exception as e:
                return False, f"Impossible d'ouvrir le fichier Excel (format corrompu ou invalide): {e}", [], {}

            ws = _get_best_sheet(wb)
            if not ws:
                return False, "Aucune feuille de calcul exploitable trouvée dans le classeur.", [], {}

            tasks = parse_tasks(file_path)
            if not tasks or len(tasks) == 0:
                return False, "Aucun point de contrôle/tâche valide détecté dans la feuille.", [], {}

            metadata = {
                "sheet_name": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "task_count": len(tasks)
            }
            return True, "Structure Excel valide.", tasks, metadata

        except Exception as e:
            return False, f"Erreur lors de la validation Excel : {e}", [], {}

    def validate_and_import_excel(
        self,
        file_path: str,
        original_filename: str,
        name: Optional[str] = None,
        checklist_code: Optional[str] = None,
        equipment_pattern: str = "ALL",
        target_type: str = "PPE",
        imported_by: str = "admin",
        change_summary: str = ""
    ) -> Tuple[bool, str, Optional[int]]:
        """
        Atomically validates, archives, and imports an Excel checklist into SQLite.
        Activates the new version and deactivates any previous version for the same definition.
        Returns: (success, message, new_version_id)
        """
        is_valid, msg, tasks, meta = self.validate_excel_file(file_path)
        if not is_valid:
            logger.warning(f"Excel validation rejected: {msg}")
            return False, f"Validation échouée: {msg}", None

        # Compute SHA-256 of uploaded file
        sha256 = ""
        with open(file_path, "rb") as f:
            sha256 = hashlib.sha256(f.read()).hexdigest()

        # Determine Code & Name
        base_name = os.path.splitext(original_filename)[0]
        if not checklist_code:
            # Extract standard code or sanitize
            match = re.search(r'(PPE-[A-Z0-9\-]+)', original_filename, re.IGNORECASE)
            if match:
                checklist_code = match.group(1).upper()
            else:
                sanitized = re.sub(r'[^A-Za-z0-9]+', '_', base_name).strip('_').upper()
                checklist_code = f"CHK_{sanitized[:24]}"

        if not name:
            name = base_name.replace('_', ' ').replace('-', ' ').strip()

        # Archive copy
        now_ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        arch_name = f"{checklist_code}_{now_ts}_{os.path.basename(original_filename)}"
        arch_path = os.path.join(self.archive_dir, arch_name)
        try:
            shutil.copy2(file_path, arch_path)
        except Exception as e:
            logger.error(f"Failed to copy file to archive: {e}")
            arch_name = os.path.basename(file_path)

        # Atomic SQLite Transaction
        with self._get_conn() as conn:
            cur = conn.cursor()
            try:
                # 1. Get or Create Definition
                cur.execute("SELECT id FROM checklist_definitions WHERE checklist_code = ?", (checklist_code,))
                def_row = cur.fetchone()
                if def_row:
                    definition_id = def_row["id"]
                    cur.execute("""
                        UPDATE checklist_definitions 
                        SET name = ?, updated_at = CURRENT_TIMESTAMP 
                        WHERE id = ?
                    """, (name, definition_id))
                else:
                    cur.execute("""
                        INSERT INTO checklist_definitions (checklist_code, name, target_type, equipment_pattern, description)
                        VALUES (?, ?, ?, ?, ?)
                    """, (checklist_code, name, target_type, equipment_pattern, f"Importé depuis {original_filename}"))
                    definition_id = cur.lastrowid

                # 2. Determine Next Version Number
                cur.execute("SELECT MAX(version_number) as max_v FROM checklist_versions WHERE definition_id = ?", (definition_id,))
                v_row = cur.fetchone()
                next_version = (v_row["max_v"] or 0) + 1

                # 3. Deactivate all existing versions for this definition
                cur.execute("""
                    UPDATE checklist_versions 
                    SET status = 'INACTIVE' 
                    WHERE definition_id = ?
                """, (definition_id,))

                # 4. Insert New ACTIVE Version
                cur.execute("""
                    INSERT INTO checklist_versions (
                        definition_id, version_number, status, original_filename, 
                        stored_filename, file_sha256, item_count, imported_by, change_summary
                    ) VALUES (?, ?, 'ACTIVE', ?, ?, ?, ?, ?, ?)
                """, (
                    definition_id, next_version, original_filename, 
                    arch_name, sha256, len(tasks), imported_by, change_summary or f"Version {next_version}"
                ))
                new_version_id = cur.lastrowid

                # 5. Insert Checklist Items
                for t in tasks:
                    cur.execute("""
                        INSERT INTO checklist_items (
                            version_id, item_number, section_header, description, method, icon, row_index
                        ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (
                        new_version_id,
                        t.get("number", 1),
                        t.get("section", ""),
                        t.get("description", ""),
                        t.get("method", "Vérification standard"),
                        t.get("icon", "fa-clipboard-check"),
                        t.get("row", 0)
                    ))

                conn.commit()
                logger.info(f"Checklist {checklist_code} V{next_version} successfully imported & activated ({len(tasks)} items).")
                return True, f"Checklist '{name}' (V{next_version}) activée avec succès avec {len(tasks)} points de contrôle.", new_version_id

            except Exception as e:
                conn.rollback()
                logger.error(f"Import rollback due to error: {e}")
                return False, f"Erreur lors de l'enregistrement en base de données : {e}", None

    # =========================================================================
    # Lifecycle & Status Management
    # =========================================================================

    def deactivate_version(self, version_id: int) -> Tuple[bool, str]:
        """
        Soft-deactivates a checklist version (sets status = 'INACTIVE').
        Historical records are completely preserved.
        """
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT v.id, v.version_number, d.name FROM checklist_versions v JOIN checklist_definitions d ON v.definition_id = d.id WHERE v.id = ?", (version_id,))
            row = cur.fetchone()
            if not row:
                return False, "Version de checklist introuvable."
            
            cur.execute("UPDATE checklist_versions SET status = 'INACTIVE' WHERE id = ?", (version_id,))
            conn.commit()
            return True, f"La checklist '{row['name']}' (V{row['version_number']}) a été désactivée de l'application active. L'historique des exécutions passées est conservé."

    def activate_version(self, version_id: int) -> Tuple[bool, str]:
        """
        Activates a specific version, deactivating other versions of the same definition.
        """
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT definition_id, version_number FROM checklist_versions WHERE id = ?", (version_id,))
            row = cur.fetchone()
            if not row:
                return False, "Version de checklist introuvable."
            
            def_id = row["definition_id"]
            # Deactivate others in same definition
            cur.execute("UPDATE checklist_versions SET status = 'INACTIVE' WHERE definition_id = ?", (def_id,))
            # Activate this one
            cur.execute("UPDATE checklist_versions SET status = 'ACTIVE' WHERE id = ?", (version_id,))
            conn.commit()
            return True, f"Version V{row['version_number']} réactivée comme version en vigueur."

    # =========================================================================
    # Queries for Runtime & Administration
    # =========================================================================

    def get_all_definitions_admin(self) -> List[Dict[str, Any]]:
        """Returns all checklist definitions and their latest/active versions for admin display."""
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT 
                    d.id as definition_id,
                    d.checklist_code,
                    d.name,
                    d.target_type,
                    d.equipment_pattern,
                    v.id as version_id,
                    v.version_number,
                    v.status,
                    v.original_filename,
                    v.stored_filename,
                    v.item_count,
                    v.imported_at,
                    v.imported_by
                FROM checklist_definitions d
                LEFT JOIN checklist_versions v ON v.definition_id = d.id
                WHERE v.id = (
                    SELECT v2.id FROM checklist_versions v2 
                    WHERE v2.definition_id = d.id 
                    ORDER BY (v2.status = 'ACTIVE') DESC, v2.version_number DESC 
                    LIMIT 1
                )
                ORDER BY (v.status = 'ACTIVE') DESC, d.name ASC
            """)
            return [dict(r) for r in cur.fetchall()]

    def get_version_history_for_definition(self, definition_id: int) -> List[Dict[str, Any]]:
        """Returns all versions for a specific definition."""
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT id, version_number, status, original_filename, stored_filename, item_count, imported_at, imported_by, change_summary
                FROM checklist_versions
                WHERE definition_id = ?
                ORDER BY version_number DESC
            """, (definition_id,))
            return [dict(r) for r in cur.fetchall()]

    def get_active_checklist(self, equipment: str = "", task_type: str = "") -> Optional[Dict[str, Any]]:
        """
        Finds the matching ACTIVE checklist version and returns its items.
        Returns dict with version info and list of task items.
        """
        with self._get_conn() as conn:
            cur = conn.cursor()
            
            # 1. Try exact or pattern match on active versions
            cur.execute("""
                SELECT 
                    v.id as version_id,
                    v.version_number,
                    v.original_filename,
                    v.stored_filename,
                    d.id as definition_id,
                    d.checklist_code,
                    d.name as checklist_name,
                    d.equipment_pattern
                FROM checklist_versions v
                JOIN checklist_definitions d ON v.definition_id = d.id
                WHERE v.status = 'ACTIVE'
                ORDER BY v.id DESC
            """)
            active_versions = [dict(r) for r in cur.fetchall()]

            if not active_versions:
                return None

            # Best match algorithm
            best_ver = None
            query = f"{equipment} {task_type}".lower().strip()
            
            for ver in active_versions:
                name_low = ver["checklist_name"].lower()
                code_low = ver["checklist_code"].lower()
                pattern_low = (ver["equipment_pattern"] or "").lower()
                
                if equipment and (equipment.lower() in name_low or equipment.lower() in pattern_low or equipment.lower() in code_low):
                    best_ver = ver
                    break
                if task_type and (task_type.lower() in name_low or task_type.lower() in pattern_low):
                    best_ver = ver
                    break

            if not best_ver and active_versions:
                best_ver = active_versions[0] # Fallback to first available active

            if not best_ver:
                return None

            # Fetch items for this version
            cur.execute("""
                SELECT id, item_number, section_header, description, method, control_type, icon, row_index
                FROM checklist_items
                WHERE version_id = ?
                ORDER BY item_number ASC
            """, (best_ver["version_id"],))
            
            items = []
            for r in cur.fetchall():
                items.append({
                    "id": r["id"],
                    "number": r["item_number"],
                    "item_number": r["item_number"],
                    "section": r["section_header"] or "",
                    "section_header": r["section_header"] or "",
                    "description": r["description"] or "",
                    "method": r["method"] or "Vérification standard",
                    "control_type": r["control_type"] or "OK_NOK_NA",
                    "icon": r["icon"] or "fa-clipboard-check",
                    "row": r["row_index"] or 0,
                    "row_index": r["row_index"] or 0
                })

            best_ver["items"] = items
            return best_ver

    def get_all_active_templates_list(self) -> List[Dict[str, Any]]:
        """Returns a list of all ACTIVE templates for dropdown selectors."""
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT 
                    v.id as version_id,
                    v.version_number,
                    v.original_filename,
                    d.checklist_code,
                    d.name as checklist_name
                FROM checklist_versions v
                JOIN checklist_definitions d ON v.definition_id = d.id
                WHERE v.status = 'ACTIVE'
                ORDER BY d.name ASC
            """)
            res = []
            for r in cur.fetchall():
                res.append({
                    "version_id": r["version_id"],
                    "filename": r["original_filename"],
                    "name": f"{r['checklist_name']} (V{r['version_number']})",
                    "checklist_code": r["checklist_code"]
                })
            return res

    def get_items_for_version(self, version_id: int) -> List[Dict[str, Any]]:
        """Returns all normalized items for a given version ID."""
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT id, item_number, section_header, description, method, control_type, icon, row_index
                FROM checklist_items
                WHERE version_id = ?
                ORDER BY item_number ASC
            """, (version_id,))
            items = []
            for r in cur.fetchall():
                items.append({
                    "id": r["id"],
                    "number": r["item_number"],
                    "item_number": r["item_number"],
                    "section": r["section_header"] or "",
                    "section_header": r["section_header"] or "",
                    "description": r["description"] or "",
                    "method": r["method"] or "Vérification standard",
                    "control_type": r["control_type"] or "OK_NOK_NA",
                    "icon": r["icon"] or "fa-clipboard-check",
                    "row": r["row_index"] or 0,
                    "row_index": r["row_index"] or 0
                })
            return items

    def get_template_file_path(
        self,
        version_id: Optional[int] = None,
        equipment: str = "",
        task_type: str = ""
    ) -> Optional[str]:
        """
        Resolves the absolute path to the master Excel template file for a given version or equipment/task.
        Preserves the master template intact for direct copy-and-fill operations.
        """
        # 1. Direct version lookup
        if version_id:
            with self._get_conn() as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT stored_filename, original_filename FROM checklist_versions WHERE id = ?",
                    (version_id,)
                )
                row = cur.fetchone()
                if row:
                    candidates = [
                        os.path.join(self.storage_dir, row["stored_filename"]),
                        os.path.join(self.storage_dir, row["original_filename"]),
                        os.path.join(self.storage_dir, "archive", row["stored_filename"]),
                    ]
                    # Check in storage directory
                    orig_clean = os.path.splitext(row["original_filename"])[0].lower()
                    if os.path.exists(self.storage_dir):
                        for f in os.listdir(self.storage_dir):
                            if f.lower().endswith(('.xlsx', '.xls', '.xlsm')) and not f.startswith('~$'):
                                if orig_clean in f.lower() or f.lower() in orig_clean:
                                    candidates.append(os.path.join(self.storage_dir, f))

                    for c in candidates:
                        if c and os.path.isfile(c):
                            return os.path.abspath(c)

        # 2. Active checklist lookup by equipment / task_type
        chk = self.get_active_checklist(equipment=equipment, task_type=task_type)
        if chk and chk.get("version_id"):
            path = self.get_template_file_path(version_id=chk["version_id"])
            if path:
                return path

        # 3. Fallback to template discovery in storage_dir
        if os.path.exists(self.storage_dir):
            from web_portal.utils.checklist_parser import find_template
            found = find_template(f"{equipment} {task_type}".strip(), self.storage_dir)
            if found and os.path.isfile(found):
                return os.path.abspath(found)

        return None

    def get_latest_execution_for_task(self, sheet: str = "", week: str = "", equipment: str = "", task_type: str = "") -> Optional[Dict[str, Any]]:
        """Finds the most recent execution record for a specific PMA task to restore answers on reload."""
        with self._get_conn() as conn:
            cur = conn.cursor()
            sql = "SELECT * FROM checklist_executions WHERE 1=1"
            params = []
            if equipment:
                sql += " AND equipment = ?"
                params.append(equipment)
            if sheet:
                sql += " AND sheet = ?"
                params.append(sheet)
            if week:
                sql += " AND week = ?"
                params.append(week)
            if task_type:
                sql += " AND task_type = ?"
                params.append(task_type)
            sql += " ORDER BY executed_at DESC LIMIT 1"
            cur.execute(sql, params)
            row = cur.fetchone()
            if row:
                d = dict(row)
                try:
                    d["answers"] = json.loads(d.get("answers_json", "{}"))
                except:
                    d["answers"] = {}
                return d
            return None

    # =========================================================================
    # Execution History
    # =========================================================================

    def record_execution(
        self,
        version_id: Optional[int],
        equipment: str,
        task_type: str,
        sheet: str,
        week: str,
        month: str,
        technician_name: str,
        technician_matricule: str,
        shift: str,
        answers: Dict[str, Any],
        filled_excel_path: str = ""
    ) -> int:
        """
        Permanently saves a checklist execution linked to its version.
        """
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO checklist_executions (
                    version_id, equipment, task_type, sheet, week, month,
                    technician_name, technician_matricule, shift, answers_json, filled_excel_path
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                version_id,
                equipment,
                task_type,
                sheet,
                week,
                month,
                technician_name,
                technician_matricule,
                shift,
                json.dumps(answers, ensure_ascii=False),
                filled_excel_path
            ))
            conn.commit()
            return cur.lastrowid

    def get_execution_history(self, equipment: Optional[str] = None, limit: int = 50) -> List[Dict[str, Any]]:
        """Returns historical completed checklist records."""
        with self._get_conn() as conn:
            cur = conn.cursor()
            if equipment:
                cur.execute("""
                    SELECT e.*, v.version_number, d.name as checklist_name
                    FROM checklist_executions e
                    LEFT JOIN checklist_versions v ON e.version_id = v.id
                    LEFT JOIN checklist_definitions d ON v.definition_id = d.id
                    WHERE e.equipment = ?
                    ORDER BY e.executed_at DESC
                    LIMIT ?
                """, (equipment, limit))
            else:
                cur.execute("""
                    SELECT e.*, v.version_number, d.name as checklist_name
                    FROM checklist_executions e
                    LEFT JOIN checklist_versions v ON e.version_id = v.id
                    LEFT JOIN checklist_definitions d ON v.definition_id = d.id
                    ORDER BY e.executed_at DESC
                    LIMIT ?
                """, (limit,))
            
            rows = []
            for r in cur.fetchall():
                d = dict(r)
                try:
                    d["answers"] = json.loads(d["answers_json"]) if d.get("answers_json") else {}
                except:
                    d["answers"] = {}
                rows.append(d)
            return rows

    def get_execution_by_id(self, exec_id: int) -> Optional[Dict[str, Any]]:
        """Returns a single execution record by its primary key ID."""
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT e.*, v.version_number, d.name as checklist_name
                FROM checklist_executions e
                LEFT JOIN checklist_versions v ON e.version_id = v.id
                LEFT JOIN checklist_definitions d ON v.definition_id = d.id
                WHERE e.id = ?
            """, (exec_id,))
            row = cur.fetchone()
            if not row:
                return None
            d = dict(row)
            try:
                d["answers"] = json.loads(d["answers_json"]) if d.get("answers_json") else {}
            except:
                d["answers"] = {}
            return d
