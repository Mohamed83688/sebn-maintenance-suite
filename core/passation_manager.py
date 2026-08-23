import os
import sqlite3
import logging
import datetime
from typing import Dict, Any, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger('sebn-maintenance')

DEFAULT_QUESTIONS = [
    ("État général de la zone / propreté du poste", "Sécurité & 5S", "CHOICE", "Conforme,Non Conforme,À Améliorer"),
    ("Vérification des équipements de sécurité et arrêts d'urgence", "Sécurité", "CHOICE", "OK,NOK,N/A"),
    ("Statut des machines critiques et lignes actives", "Production", "CHOICE", "Toutes en marche,Arrêts partiels,Panne majeure"),
    ("Pannes en cours et interventions non clôturées", "Maintenance Curative", "TEXT", ""),
    ("Pièces de rechange (PDR) consommées ou à commander", "PDR & Magasin", "TEXT", ""),
    ("Tâches préventives (PMA) réalisées durant le shift", "Maintenance Préventive", "TEXT", ""),
    ("Consignes particulières pour l'équipe montante", "Communication", "TEXT", ""),
    ("Points de blocage / Besoin d'escalade Manager", "Management", "TEXT", "")
]

class PassationManager:
    """
    Passation (Shift Handover) Management Engine.
    Handles shift handovers, questions, responses, handover logs, and Excel/PDF reports.
    """
    def __init__(self, db_path: str, data_dir: str):
        self.db_path = db_path
        self.data_dir = data_dir
        self.reports_dir = os.path.join(data_dir, "Passation_Reports")
        os.makedirs(self.reports_dir, exist_ok=True)
        self._init_db()

    def _get_conn(self):
        conn = sqlite3.connect(self.db_path, timeout=15)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_db(self):
        with self._get_conn() as conn:
            cur = conn.cursor()
            # Passation main records
            cur.execute("""
                CREATE TABLE IF NOT EXISTS passations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    technician_matricule TEXT,
                    technician_name TEXT,
                    shift TEXT NOT NULL,
                    target_shift TEXT,
                    zone_name TEXT,
                    remarks TEXT,
                    status TEXT DEFAULT 'Completed',
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Check and add missing columns for passations
            cur.execute("PRAGMA table_info(passations)")
            p_cols = {row["name"].lower() for row in cur.fetchall()}
            if "created_at" not in p_cols:
                cur.execute("ALTER TABLE passations ADD COLUMN created_at DATETIME")
                cur.execute("UPDATE passations SET created_at = COALESCE(timestamp, datetime('now')) WHERE created_at IS NULL")
            if "target_shift" not in p_cols:
                cur.execute("ALTER TABLE passations ADD COLUMN target_shift TEXT")
            if "zone_name" not in p_cols:
                cur.execute("ALTER TABLE passations ADD COLUMN zone_name TEXT")
            # Questions table
            cur.execute("""
                CREATE TABLE IF NOT EXISTS passation_questions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    text TEXT NOT NULL,
                    category TEXT NOT NULL DEFAULT 'Général',
                    type TEXT NOT NULL DEFAULT 'CHOICE',
                    options TEXT DEFAULT '',
                    is_active INTEGER DEFAULT 1,
                    sort_order INTEGER DEFAULT 0
                )
            """)

            cur.execute("PRAGMA table_info(passation_questions)")
            q_cols = {row["name"].lower() for row in cur.fetchall()}
            if "is_active" not in q_cols:
                cur.execute("ALTER TABLE passation_questions ADD COLUMN is_active INTEGER DEFAULT 1")
                if "active" in q_cols:
                    cur.execute("UPDATE passation_questions SET is_active = active WHERE is_active IS NULL")
            if "sort_order" not in q_cols:
                cur.execute("ALTER TABLE passation_questions ADD COLUMN sort_order INTEGER DEFAULT 0")
                if "display_order" in q_cols:
                    cur.execute("UPDATE passation_questions SET sort_order = display_order WHERE sort_order IS NULL")
            # Responses table
            cur.execute("""
                CREATE TABLE IF NOT EXISTS passation_responses (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    passation_id INTEGER NOT NULL,
                    question_id INTEGER NOT NULL,
                    answer TEXT,
                    FOREIGN KEY (passation_id) REFERENCES passations(id) ON DELETE CASCADE
                )
            """)
            # Passation settings
            cur.execute("""
                CREATE TABLE IF NOT EXISTS passation_settings (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL
                )
            """)

            # Seed default questions if empty
            cur.execute("SELECT COUNT(*) FROM passation_questions")
            if cur.fetchone()[0] == 0:
                for idx, (qtext, qcat, qtype, qopts) in enumerate(DEFAULT_QUESTIONS, start=1):
                    cur.execute("""
                        INSERT INTO passation_questions (text, category, type, options, is_active, sort_order)
                        VALUES (?, ?, ?, ?, 1, ?)
                    """, (qtext, qcat, qtype, qopts, idx))

            conn.commit()

    # ── Questions Management ─────────────────────────────────────────────────

    def get_questions(self, active_only: bool = True) -> List[Dict[str, Any]]:
        with self._get_conn() as conn:
            cur = conn.cursor()
            if active_only:
                cur.execute("SELECT * FROM passation_questions WHERE is_active = 1 ORDER BY sort_order ASC, id ASC")
            else:
                cur.execute("SELECT * FROM passation_questions ORDER BY sort_order ASC, id ASC")
            return [dict(r) for r in cur.fetchall()]

    def add_question(self, text: str, category: str, qtype: str = "CHOICE", options: str = "") -> int:
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO passation_questions (text, category, type, options, is_active, sort_order)
                VALUES (?, ?, ?, ?, 1, (SELECT COALESCE(MAX(sort_order), 0) + 1 FROM passation_questions))
            """, (text.strip(), category.strip(), qtype.strip(), options.strip()))
            conn.commit()
            return cur.lastrowid

    def get_question(self, qid: int) -> Optional[Dict[str, Any]]:
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM passation_questions WHERE id = ?", (qid,))
            row = cur.fetchone()
            return dict(row) if row else None

    def update_question(self, qid: int, text: str, category: str, qtype: str = "CHOICE", options: str = "", sort_order: int = 1) -> bool:
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                UPDATE passation_questions
                SET text = ?, category = ?, type = ?, options = ?, sort_order = ?
                WHERE id = ?
            """, (text.strip(), category.strip(), qtype.strip(), options.strip(), sort_order, qid))
            conn.commit()
            return cur.rowcount > 0

    def delete_question(self, qid: int) -> bool:
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM passation_questions WHERE id = ?", (qid,))
            conn.commit()
            return cur.rowcount > 0

    def toggle_question(self, qid: int) -> Tuple[bool, int]:
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT is_active FROM passation_questions WHERE id = ?", (qid,))
            row = cur.fetchone()
            if not row: return False, 0
            new_st = 0 if row["is_active"] == 1 else 1
            cur.execute("UPDATE passation_questions SET is_active = ? WHERE id = ?", (new_st, qid))
            conn.commit()
            return True, new_st

    # ── Passation Records ────────────────────────────────────────────────────

    def create_passation(
        self,
        user_id: Optional[int],
        technician_name: str,
        technician_matricule: str,
        shift: str,
        target_shift: str,
        zone_name: str,
        remarks: str,
        answers: Dict[int, str]
    ) -> int:
        """Saves a shift handover and records all questionnaire responses atomically."""
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO passations (
                    user_id, technician_matricule, technician_name, shift, target_shift, zone_name, remarks, status, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, 'Completed', CURRENT_TIMESTAMP)
            """, (
                user_id,
                technician_matricule.strip(),
                technician_name.strip(),
                shift.strip(),
                target_shift.strip(),
                zone_name.strip(),
                remarks.strip()
            ))
            pass_id = cur.lastrowid

            for qid, ans in answers.items():
                cur.execute("""
                    INSERT INTO passation_responses (passation_id, question_id, answer)
                    VALUES (?, ?, ?)
                """, (pass_id, int(qid), str(ans).strip()))

            conn.commit()
            return pass_id

    def get_passations(self, limit: int = 50) -> List[Dict[str, Any]]:
        """Retrieves recent shift handovers with response count."""
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT p.*, COUNT(r.id) as response_count
                FROM passations p
                LEFT JOIN passation_responses r ON p.id = r.passation_id
                GROUP BY p.id
                ORDER BY p.created_at DESC
                LIMIT ?
            """, (limit,))
            return [dict(r) for r in cur.fetchall()]

    def get_passation_detail(self, pass_id: int) -> Optional[Dict[str, Any]]:
        """Retrieves full detail of a specific handover including all questions and answers."""
        with self._get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM passations WHERE id = ?", (pass_id,))
            p_row = cur.fetchone()
            if not p_row:
                return None

            pass_dict = dict(p_row)
            cur.execute("""
                SELECT q.id as question_id, q.text, q.category, q.type, r.answer
                FROM passation_questions q
                LEFT JOIN passation_responses r ON q.id = r.question_id AND r.passation_id = ?
                ORDER BY q.sort_order ASC, q.id ASC
            """, (pass_id,))
            pass_dict["responses"] = [dict(r) for r in cur.fetchall()]
            return pass_dict

    def export_passation_excel(self, pass_id: int) -> Optional[str]:
        """Generates a professional Excel report for a shift handover."""
        detail = self.get_passation_detail(pass_id)
        if not detail:
            return None

        import re
        safe_shift = re.sub(r'[^a-zA-Z0-9_-]+', '_', detail.get('shift', 'SHIFT')).strip('_')
        safe_date = re.sub(r'[^a-zA-Z0-9_-]+', '_', (detail.get('created_at') or 'date')[:10]).strip('_')
        out_name = f"Passation_{safe_shift}_{safe_date}_{pass_id}.xlsx"
        out_path = os.path.join(self.reports_dir, out_name)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Journal de Passation"

        thin = Side(border_style='thin', color='CBD5E1')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header Title
        ws['A1'] = "RAPPORT DE PASSATION DE CONSIGNES — SEBN-TN"
        ws['A1'].font = Font(bold=True, size=14, color='1E3A8A')
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws['A2'] = f"Date: {detail['created_at']}   |   Shift Sortant: {detail['shift']}   →   Shift Montant: {detail.get('target_shift', '')}"
        ws.merge_cells('A2:E2')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(italic=True, size=10, color='64748B')

        ws['A3'] = f"Technicien / Responsable: {detail['technician_name']} ({detail['technician_matricule']})   |   Zone: {detail.get('zone_name', 'Toutes')}"
        ws.merge_cells('A3:E3')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A3'].font = Font(size=10, color='64748B')

        ws.append([])

        # Table Header
        headers = ['#', 'Catégorie', 'Point de Contrôle / Consigne', 'Réponse / État']
        ws.append(headers)
        hdr_row = ws.max_row
        for ci in range(1, len(headers) + 1):
            c = ws.cell(row=hdr_row, column=ci)
            c.font = Font(bold=True, size=11, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor='1E3A8A')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border

        for idx, resp in enumerate(detail.get("responses", []), start=1):
            ans = resp.get("answer") or "–"
            ws.append([idx, resp["category"], resp["text"], ans])
            ri = ws.max_row
            for ci in range(1, 5):
                c = ws.cell(row=ri, column=ci)
                c.border = border
                c.alignment = Alignment(vertical='center', wrap_text=(ci in (3, 4)))

        # Remarks Block
        if detail.get("remarks"):
            ws.append([])
            ws.append(['', 'Remarques Générales / Pannes Transmises:', '', ''])
            ws.merge_cells(f'B{ws.max_row}:D{ws.max_row}')
            ws.cell(row=ws.max_row, column=2).font = Font(bold=True, color='1E3A8A')

            ws.append(['', detail["remarks"], '', ''])
            ws.merge_cells(f'B{ws.max_row}:D{ws.max_row}')
            ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True)

        for ci, w in enumerate([6, 20, 50, 35], start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        wb.save(out_path)
        return out_path
