"""
checklist_parser.py
====================
Web-safe parser for PPE (Preventive Maintenance Checklist) Excel templates.
Ported from ui/checklist_wizard.py (desktop) with win32com dependencies removed.

Key design principles:
  1. Load a COPY of the original template (preserve structure, styles, layout).
  2. Auto-detect which column holds this month's / this week's results.
  3. Auto-detect which column holds observations/remarks.
  4. Fill metadata (machine name, technician, date) in the header area.
  5. Write OK/NOK/N/A values and observations directly into the template cells.
  6. Save the filled copy to PPE_Filled/ — same structure as the manager's template.
"""

import os
import re
import shutil
import difflib
import datetime
import unicodedata

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.cell.cell import MergedCell
    import openpyxl.utils  # for get_column_letter in save_filled_checklist logging
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _norm(t) -> str:
    """Unicode-normalize + uppercase for robust comparison (matches desktop _norm)."""
    if not t:
        return ""
    return "".join(
        c for c in unicodedata.normalize('NFD', str(t))
        if unicodedata.category(c) != 'Mn'
    ).upper().strip()


def _get_writeable_cell(ws, cell):
    """
    If `cell` is a MergedCell, return the top-left master cell.
    Otherwise return the cell itself.
    """
    if not isinstance(cell, MergedCell):
        return cell
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return None


def _get_adjacent_empty_cell(ws, r_or_cell, c=None):
    """
    Finds the first empty writeable cell immediately to the right or below a label cell.
    Correctly accounts for vertical and horizontal MergedCell ranges.
    """
    if c is None:
        r = r_or_cell.row
        c = r_or_cell.column
    else:
        r = r_or_cell

    cell = ws.cell(row=r, column=c)
    max_c = c
    max_r = r
    min_r = r
    min_c = c

    if isinstance(cell, MergedCell) or hasattr(ws, 'merged_cells'):
        for rng in ws.merged_cells.ranges:
            if cell.coordinate in rng:
                max_c = rng.max_col
                max_r = rng.max_row
                min_r = rng.min_row
                min_c = rng.min_col
                break

    # 1. Check cell immediately to the right of merged range
    cand_right = _get_writeable_cell(ws, ws.cell(row=min_r, column=max_c + 1))
    if cand_right and (cand_right.value is None or str(cand_right.value).strip() == ''):
        return cand_right

    # 2. Check cell immediately below merged range
    cand_below = _get_writeable_cell(ws, ws.cell(row=max_r + 1, column=min_c))
    if cand_below and (cand_below.value is None or str(cand_below.value).strip() == ''):
        return cand_below

    return None


def _get_best_sheet(wb):
    """
    Find the best sheet in the workbook (preferring visible non-empty data sheets).
    Skips: hidden sheets, 'OLD_' prefixed sheets, 'Feuil*' utility sheets,
    and 'Signatures' sheets.
    Falls back to largest sheet if no clean candidate found.
    """
    SKIP_PREFIXES  = ('OLD_', 'old_', 'ANCIEN', 'BACKUP', 'FEUIL', 'Feuil', 'SIGNATURES', 'PARAM')
    SKIP_SUBSTRINGS = ('param', 'PARAM', 'old', 'OLD', 'backup', 'BACKUP')

    candidates = []
    for sname in wb.sheetnames:
        s = wb[sname]
        if s.sheet_state == 'hidden':
            continue
        su = sname.upper()
        if any(su.startswith(pfx.upper()) for pfx in SKIP_PREFIXES):
            continue
        candidates.append(s)

    if not candidates:
        # If all filtered out, return first visible sheet
        for sname in wb.sheetnames:
            s = wb[sname]
            if s.sheet_state != 'hidden':
                return s
        return wb.active

    # Among candidates, pick the one with most filled cells (not just dimensions)
    best_sheet = None
    max_score  = -1
    for s in candidates:
        # Score = number of cells that are actually non-empty (sample first 50 rows)
        score = 0
        for row in s.iter_rows(min_row=1, max_row=min(50, s.max_row), values_only=True):
            score += sum(1 for c in row if c is not None and str(c).strip())
        if score > max_score:
            max_score  = score
            best_sheet = s

    return best_sheet or wb.active


# ──────────────────────────────────────────────────────────────────────────────
# Template Discovery
# ──────────────────────────────────────────────────────────────────────────────

def find_template(type_name: str, templates_dir: str):
    """
    Find the best-matching PPE template Excel file for a given task type / machine.
    Uses Jaccard token overlap + difflib sequence ratio.
    If no threshold match (> 0.15) is found but templates exist, falls back to the first available template.
    """
    if not templates_dir or not os.path.isdir(templates_dir):
        return None

    candidates = [
        f for f in os.listdir(templates_dir)
        if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$')
    ]
    if not candidates:
        return None

    query = str(type_name).lower().strip()
    query_tokens = set(re.split(r'[\s\-_/]+', query))

    best_path  = None
    best_score = 0.0

    for fname in candidates:
        name   = os.path.splitext(fname)[0].lower()
        tokens = set(re.split(r'[\s\-_/]+', name))

        overlap = len(query_tokens & tokens) / max(len(query_tokens | tokens), 1)
        ratio   = difflib.SequenceMatcher(None, query, name).ratio()
        score   = overlap * 0.6 + ratio * 0.4

        if score > best_score:
            best_score = score
            best_path  = os.path.join(templates_dir, fname)

    if best_score > 0.15 and best_path:
        return best_path

    # Fallback: return the first candidate template if any exist
    return os.path.join(templates_dir, candidates[0])


def list_all_templates(templates_dir: str) -> list:
    """
    Return a list of dicts for all available PPE templates in templates_dir:
    [{'filename': 'test.xlsx', 'name': 'test', 'path': '...'}, ...]
    """
    if not templates_dir or not os.path.isdir(templates_dir):
        return []
    candidates = [
        f for f in os.listdir(templates_dir)
        if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$')
    ]
    templates = []
    for f in sorted(candidates):
        templates.append({
            'filename': f,
            'name': os.path.splitext(f)[0],
            'path': os.path.join(templates_dir, f)
        })
    return templates


# ──────────────────────────────────────────────────────────────────────────────
# Column Detection
# ──────────────────────────────────────────────────────────────────────────────

# ──────────────────────────────────────────────────────────────────────────────
# Column Detection
# ──────────────────────────────────────────────────────────────────────────────

def _detect_result_obs_columns(ws, week_str: str = ""):
    """
    Detect which column should receive the OK/NOK result value (col_target)
    and which should receive the observation remark (col_obs).
    Returns (col_target: int, col_obs: int) — 1-based column numbers.
    """
    now           = datetime.datetime.now()
    current_month = now.month
    iso_week      = now.isocalendar()[1]

    col_target = None
    col_obs    = None
    max_col    = min(ws.max_column + 1, 100)

    # ── S0: Direct EXACT match on week_str (e.g. "S33" or "KW 33" or "33") ──────
    if week_str:
        norm_w = _norm(week_str)
        for r in range(1, 30):
            for c in range(5, max_col):
                val = _norm(ws.cell(row=r, column=c).value)
                if val and (val == norm_w or val == f"S{norm_w}" or val == f"W{norm_w}" or val == f"KW{norm_w}"):
                    col_target = c
                    break
            if col_target:
                break

    # ── S1: EXACT match for current ISO week (e.g. "S33", "KW33", "W33") ───────
    if not col_target:
        exact_keys = [f"S{iso_week}", f"W{iso_week}", f"KW{iso_week}", f"KW {iso_week}", f"WEEK {iso_week}"]
        for r in range(1, 30):
            for c in range(5, max_col):
                val = _norm(ws.cell(row=r, column=c).value)
                if val and val in exact_keys:
                    col_target = c
                    break
            if col_target:
                break

    # ── S2: Contiguous 1-12 monthly grid (EXACT matching) ──────────────────────
    if not col_target:
        for r in range(1, 30):
            for c in range(5, max(5, max_col - 10)):
                vals = []
                for k in range(12):
                    try:
                        vals.append(int(str(ws.cell(row=r, column=c+k).value or '').strip()))
                    except Exception:
                        break
                if vals == list(range(1, 13)):
                    col_target = c + (current_month - 1)
                    break
            if col_target:
                break

    # ── S3: Month name exact match ──────────────────────────────────────────────
    if not col_target:
        months_fr_short = ["JAN","FEV","MAR","AVR","MAI","JUI","JUL","AOU","SEP","OCT","NOV","DEC"]
        months_fr_long  = ["JANVIER","FEVRIER","MARS","AVRIL","MAI","JUIN","JUILLET",
                           "AOUT","SEPTEMBRE","OCTOBRE","NOVEMBRE","DECEMBRE"]
        month_idx = max(0, min(11, current_month - 1))
        m_short   = months_fr_short[month_idx]
        m_long    = months_fr_long[month_idx]
        for r in range(1, 30):
            for c in range(5, max_col):
                val = _norm(ws.cell(row=r, column=c).value)
                if val and (val == m_short or val == m_long or m_short in val or m_long in val):
                    col_target = c
                    break
            if col_target:
                break

    # ── S4: Result/Check keyword in header ────────────────────────────────────
    if not col_target:
        result_kw = ["CONTRÔLE","CONTROLE","CHECK","RÉSULTAT","RESULTAT",
                     "VALEUR","VALUE","MESURE","MEASURE"]
        for r in range(1, 20):
            for c in range(7, max_col):
                val = str(ws.cell(row=r, column=c).value or "").strip().upper()
                if any(k in val for k in result_kw):
                    col_target = c
                    break

    # ── S5: Fallback to column 8 (Column H) ───────────────────────────────────
    if not col_target:
        col_target = 8

    col_obs = col_target + 1
    return col_target, col_obs


# ──────────────────────────────────────────────────────────────────────────────
# Task Row Extraction
# ──────────────────────────────────────────────────────────────────────────────

_CTRL_KEYWORDS = {'mode', 'ctrl', 'control', 'contrôle', 'type', 'méthode', 'method'}

def _is_section_header(val: str) -> bool:
    v = val.strip().lower()
    if len(v) > 120:
        return False
    if val.strip().isupper() and len(val.strip()) > 3:
        return True
    if re.match(r'^[IVXivx]{1,4}\s*[\.\-–]|^\d+[\.\-–]\s|^[A-Z]\s*[\.\-–]', val.strip()):
        return True
    return False


def _detect_description_col(header_row: list):
    keywords = ['tâche', 'description', 'opération', 'vérification', 'point',
                'task', 'operation', 'activité', 'contrôle', 'libellé', 'inspection']
    for i, cell in enumerate(header_row):
        val = str(cell).lower().strip() if cell else ''
        if any(kw in val for kw in keywords):
            return i
    return None


def _detect_method_col(header_row: list):
    for i, cell in enumerate(header_row):
        val = str(cell).lower().strip() if cell else ''
        if any(kw in val for kw in _CTRL_KEYWORDS):
            return i
    return None


def _guess_icon(description: str) -> str:
    d = description.lower()
    if any(w in d for w in ['nettoy', 'souffl', 'clean']):         return 'fa-broom'
    if any(w in d for w in ['visuel', 'inspect', 'contrôl', 'vérif']): return 'fa-eye'
    if any(w in d for w in ['électr', 'câbl', 'connex', 'circuit']): return 'fa-plug'
    if any(w in d for w in ['pression', 'fluide', 'niveau', 'manomet']): return 'fa-gauge'
    if any(w in d for w in ['sécur', 'arrêt', 'urgence', 'protec']): return 'fa-shield-halved'
    if any(w in d for w in ['courr', 'chaîne', 'organe', 'mobile']): return 'fa-link'
    if any(w in d for w in ['lubr', 'graiss', 'huile']):            return 'fa-oil-can'
    if any(w in d for w in ['test', 'fonct', 'démarr', 'essai']):   return 'fa-play-circle'
    if any(w in d for w in ['mesur', 'valeur', 'relev']):           return 'fa-ruler'
    if any(w in d for w in ['serr', 'boulon', 'vis', 'fixation']):  return 'fa-screwdriver-wrench'
    return 'fa-clipboard-check'


def _detect_task_rows(ws, col_target: int) -> list:
    tasks = []

    # Number column detection (seq 1, 2, 3...)
    num_col = 1
    best_score = 0
    for c in range(1, 10):
        score = 0
        last  = 0
        for r in range(1, 200):
            raw = str(ws.cell(row=r, column=c).value or "").strip()
            digits = "".join(filter(str.isdigit, raw))
            if digits:
                v = int(digits)
                if v == last + 1:
                    score += 10
                    last = v
                elif 0 < v < 500:
                    score += 1
        if score > best_score:
            best_score = score
            num_col    = c

    # Header row detection
    header_idx = 0
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True)):
        filled = sum(1 for c in row if c is not None and str(c).strip())
        if filled >= 3:
            header_idx = i + 1
            break

    if header_idx == 0:
        return []  # Can't determine header row

    max_scan_col = max(2, min(25, ws.max_column + 1))
    header_row   = [ws.cell(row=header_idx, column=c).value for c in range(1, max_scan_col)]
    desc_col     = _detect_description_col(header_row)
    method_col   = _detect_method_col(header_row)

    if desc_col is None:
        desc_col = num_col

    task_num = 0
    for r in range(header_idx + 1, ws.max_row + 1):
        num_val = str(ws.cell(row=r, column=num_col).value or "").strip()
        digits  = "".join(filter(str.isdigit, num_val))
        if not digits:
            continue
        try:
            row_num = int(digits)
        except ValueError:
            continue

        desc_ws_col = desc_col + 1 if desc_col is not None else num_col + 1
        desc_val = str(ws.cell(row=r, column=desc_ws_col).value or "").strip()
        if not desc_val or desc_val.lower() in ('none', 'nan', ''):
            continue
        if len(desc_val) < 5:
            continue
        if _is_section_header(desc_val):
            continue

        method_val = ''
        if method_col is not None:
            method_ws_col = method_col + 1
            method_val = str(ws.cell(row=r, column=method_ws_col).value or "").strip()
            if method_val.lower() in ('none', 'nan'):
                method_val = ''

        task_num += 1
        tasks.append({
            'number':      task_num,
            'row':         r,
            'description': desc_val,
            'method':      method_val or 'Vérification standard',
            'icon':        _guess_icon(desc_val),
        })

        if task_num >= 60:
            break

    return tasks


def parse_tasks(template_path: str) -> list:
    """
    Extract inspection tasks from a PPE template Excel file.
    Returns list of dicts: [{number, row, description, method, icon}, ...]
    Supports .xlsx, .xlsm, and misnamed .xls files (opened as binary stream).
    """
    if not OPENPYXL_OK:
        return []
    try:
        # Try opening as binary stream — handles .xls files that are actually .xlsx
        with open(template_path, 'rb') as fh:
            wb = load_workbook(fh, data_only=True)
        ws = _get_best_sheet(wb)
        col_t, _ = _detect_result_obs_columns(ws)
        return _detect_task_rows(ws, col_t)
    except Exception as e:
        print(f"[checklist_parser] Error parsing {template_path}: {e}")
        return []


# ──────────────────────────────────────────────────────────────────────────────
# Metadata fill
# ──────────────────────────────────────────────────────────────────────────────

def _fill_metadata(ws, metadata: dict, mission_dt=None, col_target=None):
    now = mission_dt or datetime.datetime.now()
    months_fr = ["Janvier","Février","Mars","Avril","Mai","Juin",
                 "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
    month_name = months_fr[max(0, min(11, now.month - 1))]
    iso_week   = now.isocalendar()[1]
    date_fr    = now.strftime("%d/%m/%Y")

    tech  = metadata.get('technician', '')
    mat   = metadata.get('matricule', '')
    equip = metadata.get('equip', '')
    shift = metadata.get('shift', '')

    val_tech = f"{tech} ({mat})" if mat else tech

    filled_items = set()

    # Pass 1: Header area ONLY (Rows 1 to 7) to avoid touching task descriptions or task table rows
    for row_idx in range(1, 8):
        for col_idx in range(1, min(31, ws.max_column + 1)):
            cell     = ws.cell(row=row_idx, column=col_idx)
            raw_val  = str(cell.value or "").strip()
            cell_key = _norm(raw_val)
            if not cell_key:
                continue

            # Fill Machine / Equipment
            if 'equip' not in filled_items and equip:
                if any(k in cell_key for k in ['MACHINE', 'EQUIPEMENT', 'EQUIPMENT', 'TEST N']):
                    target = _get_writeable_cell(ws, cell)
                    if target:
                        if raw_val.endswith(':'):
                            target.value = raw_val + " " + equip
                        elif '..' in raw_val or '__' in raw_val:
                            sep = '..' if '..' in raw_val else '__'
                            target.value = raw_val.split(sep)[0].strip() + " " + equip
                        elif not target.value or cell_key in ['MACHINE- NO', 'MACHINE NO', 'EQUIPEMENT']:
                            target.value = equip
                        filled_items.add('equip')
                        continue

            # Fill Month
            if 'month' not in filled_items:
                if cell_key in ['MOIS', 'MONTH', 'MOIS:']:
                    target = _get_writeable_cell(ws, cell)
                    if target:
                        if raw_val.endswith(':'):
                            target.value = raw_val + " " + month_name
                        else:
                            adj = _get_writeable_cell(ws, ws.cell(row=row_idx, column=col_idx+1))
                            if adj and not adj.value:
                                adj.value = month_name
                            else:
                                target.value = month_name
                        filled_items.add('month')
                        continue

            # Fill Date
            if 'date' not in filled_items:
                if cell_key in ['DATE', 'DATE:']:
                    target = _get_writeable_cell(ws, cell)
                    if target:
                        if raw_val.endswith(':'):
                            target.value = raw_val + " " + date_fr
                        else:
                            adj = _get_writeable_cell(ws, ws.cell(row=row_idx, column=col_idx+1))
                            if adj and not adj.value:
                                adj.value = date_fr
                            else:
                                target.value = date_fr
                        filled_items.add('date')
                        continue

            # Fill Technician / Matricule in header if present
            if 'tech' not in filled_items and tech:
                if any(k in cell_key for k in ['TECHNICIEN', 'TECHNICIAN', 'NOM & PRENOM', 'NOM ET PRENOM']):
                    target = _get_writeable_cell(ws, cell)
                    if target:
                        if raw_val.endswith(':'):
                            target.value = raw_val + " " + val_tech
                        else:
                            target.value = val_tech
                        filled_items.add('tech')
                        continue

    # Pass 2: Signature blocks filling inside the table
    sig_col = col_target or 8
    for r in range(8, ws.max_row + 1):
        for c in range(1, min(10, ws.max_column + 1)):
            raw = str(ws.cell(row=r, column=c).value or "").strip()
            nv  = _norm(raw)
            if not nv:
                continue

            # Technician signature label
            if any(kw in nv for kw in ['SIGNATURE DU TECHNICIEN', 'SIGNATURE TECHNICIEN', 'NOM DU TECHNICIEN', 'TECHNICIEN / MATRICULE', 'TECHNICIEN/MATRICULE']):
                target_cell = ws.cell(row=r, column=sig_col)
                target = _get_writeable_cell(ws, target_cell) or _get_adjacent_empty_cell(ws, r, c)
                if target:
                    target.value = f"{val_tech} — {date_fr}"

            # Chef d'équipe signature label (EXCLUDE 'INFORMÉ')
            elif ('SIGNATURE CHEF' in nv or 'DATE, SIGNATURE CHEF' in nv) and 'INFORM' not in nv:
                target_cell = ws.cell(row=r+1, column=sig_col) if r + 1 <= ws.max_row else ws.cell(row=r, column=sig_col)
                target = _get_writeable_cell(ws, target_cell) or _get_adjacent_empty_cell(ws, r, c)
                if target:
                    target.value = f"{date_fr} — Validé (Équipe {shift})"


# ──────────────────────────────────────────────────────────────────────────────
# Fallback Generic Checklist
# ──────────────────────────────────────────────────────────────────────────────

GENERIC_TASKS = [
    {'number': 1, 'row': None, 'description': "Vérification visuelle générale de l'état de la machine", 'method': 'Inspection visuelle',         'icon': 'fa-eye'},
    {'number': 2, 'row': None, 'description': "Nettoyage des surfaces, filtres et ventilations",         'method': 'Nettoyage & Soufflage',        'icon': 'fa-broom'},
    {'number': 3, 'row': None, 'description': "Contrôle des connexions électriques et câblages",         'method': 'Serrage & Inspection',         'icon': 'fa-plug'},
    {'number': 4, 'row': None, 'description': "Vérification des niveaux de pression / fluides",          'method': 'Lecture manomètre',            'icon': 'fa-gauge'},
    {'number': 5, 'row': None, 'description': "Contrôle des équipements de sécurité et arrêts d'urgence",'method': 'Déclenchement & Reset',        'icon': 'fa-shield-halved'},
    {'number': 6, 'row': None, 'description': "Inspection des courroies, chaînes et organes mobiles",    'method': 'Contrôle visuel & tactile',    'icon': 'fa-link'},
    {'number': 7, 'row': None, 'description': "Lubrification des points de graissage",                   'method': 'Application graisse',          'icon': 'fa-oil-can'},
    {'number': 8, 'row': None, 'description': "Test de fonctionnement à vide",                           'method': 'Démarrage & écoute',           'icon': 'fa-play-circle'},
]


# ──────────────────────────────────────────────────────────────────────────────
# Save Filled Checklist — TEMPLATE-PRESERVING
# ──────────────────────────────────────────────────────────────────────────────

def save_filled_checklist(
    tasks:       list,
    answers:     dict,
    metadata:    dict,
    output_path: str
) -> bool:
    """
    Fill and save a completed checklist by COPYING the original master template
    and writing technician answers into the existing cells of that template.

    Rules enforced:
    - ALWAYS uses the master template as source (shutil.copy2)
    - NEVER creates a new workbook or new columns
    - NEVER modifies the master template itself
    - Writes only into pre-existing cells at the exact Excel rows detected
    - Unanswered questions → cells remain empty
    - Answered questions  → +, -, or N/A written to the correct cell
    - Values / observations → written to the detected observation column cell
    - Strictly verifies that output_path exists and is non-empty on disk
    """
    if not OPENPYXL_OK:
        print("[EXCEL] ERROR: openpyxl is not available — cannot save checklist.")
        return False

    template_path = metadata.get('template_path', '')
    print(f"[EXCEL] Template path: {template_path}")
    print(f"[EXCEL] Completed file path: {output_path}")

    # If template_path not provided or doesn't exist, search for existing template in ppe directory
    if not template_path or not os.path.isfile(template_path):
        equip = metadata.get('equip', '')
        type_ = metadata.get('type', '')
        # Try finding template in default locations
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        ppe_dir = os.path.join(base_dir, 'data', 'PPE_Templates')
        if os.path.isdir(ppe_dir):
            found = find_template(f"{equip} {type_}".strip(), ppe_dir)
            if found and os.path.isfile(found):
                template_path = found
                print(f"[EXCEL] Auto-discovered template path: {template_path}")

    template_exists = bool(template_path and os.path.isfile(template_path))
    print(f"[EXCEL] Template exists: {template_exists}")
    if not template_exists:
        print(f"[EXCEL] ERROR: Master template file not found: '{template_path}'")
        return False

    out_dir = os.path.dirname(os.path.abspath(output_path))
    os.makedirs(out_dir, exist_ok=True)
    print(f"[EXCEL] Output directory exists: {os.path.isdir(out_dir)}")

    # ── Step 1: Copy master template (preserves 100% of structure/design) ─────
    shutil.copy2(template_path, output_path)
    print(f"[EXCEL] Master template copied to: {output_path}")

    # ── Step 2: Open the COPY (never open master for writing) ─────────────────
    try:
        wb = load_workbook(output_path, keep_vba=False)
        print(f"[EXCEL] Workbook loaded: True")
    except Exception as e:
        print(f"[EXCEL] ERROR: Cannot open copied workbook: {e}")
        return False

    ws = _get_best_sheet(wb)
    print(f"[EXCEL] Active worksheet: '{ws.title}'")

    # ── Step 3: Detect which column receives Result and Observation ────────────
    week_str = metadata.get('week', '')
    col_target, col_obs = _detect_result_obs_columns(ws, week_str)
    print(f"[EXCEL] Result column: {col_target} ({openpyxl.utils.get_column_letter(col_target)}), "
          f"Observation column: {col_obs} ({openpyxl.utils.get_column_letter(col_obs)})")

    color_ok  = "008000"   # green  → +
    color_nok = "FF0000"   # red    → -
    color_na  = "0070C0"   # blue   → N/A

    # ── Step 4: Write technician answers into existing cells only ─────────────
    print(f"[EXCEL] Filling cells...")
    written = 0
    for task in tasks:
        row_num  = task.get('row') or task.get('row_index')
        task_num = task.get('number') or task.get('item_number')

        if not row_num:
            continue

        ans    = answers.get(str(task_num), {})
        status = ans.get('status', '').strip().upper()
        val    = ans.get('val', '').strip()
        obs    = ans.get('obs', '').strip()

        # If technician left this question blank → leave cell empty
        if not status and not val and not obs:
            continue

        # Map web status → template symbol
        symbol = (
            '+'   if status in ('OK',  '+')          else
            '-'   if status in ('NOK', '-')           else
            'N/A' if status in ('N/A', 'NA')          else
            ''
        )

        if symbol:
            cell = ws.cell(row=row_num, column=col_target)
            dest = _get_writeable_cell(ws, cell)
            if dest:
                dest.value = symbol
                color = color_ok if symbol == '+' else (color_nok if symbol == '-' else color_na)
                try:
                    dest.font = Font(bold=True, color=color, name="Calibri", size=11)
                except Exception:
                    dest.font = Font(bold=True, color=color)
                written += 1

        # Observation / value → adjacent observation cell
        obs_parts = []
        if val:
            obs_parts.append(val)
        if obs:
            obs_parts.append(obs)
        full_obs = " | ".join(obs_parts)

        if full_obs:
            obs_cell = ws.cell(row=row_num, column=col_obs)
            obs_dest = _get_writeable_cell(ws, obs_cell)
            if obs_dest:
                obs_dest.value = full_obs
                try:
                    obs_dest.font = Font(
                        color=color_nok if symbol == '-' else "000000",
                        name="Calibri", size=10
                    )
                except Exception:
                    pass

    # ── Step 5: Fill header metadata into existing header cells ───────────────
    try:
        _fill_metadata(ws, metadata, datetime.datetime.now(), col_target=col_target)
    except Exception as e:
        print(f"[EXCEL] Metadata fill warning: {e}")

    # ── Step 6: Save the COPY and strictly verify physical existence ──────────
    print(f"[EXCEL] Saving...")
    wb.save(output_path)
    wb.close()

    # Physical verification
    file_exists = os.path.isfile(output_path)
    file_size = os.path.getsize(output_path) if file_exists else 0
    print(f"[EXCEL] Saved successfully")
    print(f"[EXCEL] Exists: {file_exists}")
    print(f"[EXCEL] Size: {file_size} bytes")

    if not file_exists or file_size == 0:
        print(f"[EXCEL] ERROR: Physical file validation failed for {output_path}")
        return False

    return True
