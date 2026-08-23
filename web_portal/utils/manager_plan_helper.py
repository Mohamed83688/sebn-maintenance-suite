"""
manager_plan_helper.py
========================
Read and write the Manager Action Plan Excel workbook
(manager_plan_performance.xlsx) for the web portal.

Sheet layout (mirrors desktop ManagerPlanDashboard):
  - "Réunions"     : Columns [Date, Titre, Responsable, Participants, Statut, Notes]
  - "Plan d'Action": Columns [ID, Action, Responsable, Échéance, Statut, Notes, Réunion_Ref]
"""

import os
import uuid
import datetime

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# Column definitions
MEETING_COLS  = ['Date', 'Titre', 'Responsable', 'Participants', 'Statut', 'Notes']
ACTION_COLS   = ['ID', 'Action', 'Responsable', 'Échéance', 'Statut', 'Notes', 'Réunion_Ref']

VALID_MEETING_STATUSES = ['Planifiée', 'En cours', 'Terminée', 'Annulée']
VALID_ACTION_STATUSES  = ['Backlog', 'En cours', 'Terminée', 'Bloquée']


def _get_workbook(path: str) -> 'openpyxl.Workbook':
    """Load an existing workbook or create a fresh one with the correct sheets."""
    if os.path.exists(path):
        try:
            wb = load_workbook(path)
        except Exception:
            wb = _create_fresh_workbook()
    else:
        wb = _create_fresh_workbook()
    return wb


def _create_fresh_workbook() -> 'openpyxl.Workbook':
    wb = Workbook()

    thin = Side(border_style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', fgColor='2563EB')
    hdr_align = Alignment(horizontal='center', vertical='center')

    def _write_header(ws, cols):
        ws.append(cols)
        for cell in ws[1]:
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = hdr_align
            cell.border = border

    # Sheet 1 — Réunions
    ws1 = wb.active
    ws1.title = 'Réunions'
    _write_header(ws1, MEETING_COLS)

    # Sheet 2 — Plan d'Action
    ws2 = wb.create_sheet('Plan d\'Action')
    _write_header(ws2, ACTION_COLS)

    return wb


def _ensure_sheets(wb: 'openpyxl.Workbook') -> tuple:
    """Return (ws_meetings, ws_actions), creating sheets if missing."""
    sheet_names = wb.sheetnames

    if 'Réunions' not in sheet_names:
        ws1 = wb.create_sheet('Réunions')
        ws1.append(MEETING_COLS)
    else:
        ws1 = wb['Réunions']

    action_sheet = "Plan d'Action"
    if action_sheet not in sheet_names:
        ws2 = wb.create_sheet(action_sheet)
        ws2.append(ACTION_COLS)
    else:
        ws2 = wb[action_sheet]

    return ws1, ws2


# ──────────────────────────────────────────────────────────────────────────────
# Meetings
# ──────────────────────────────────────────────────────────────────────────────

def load_meetings(path: str) -> list:
    """
    Load all meetings from the Excel file.
    Returns a list of dicts with keys: Date, Titre, Responsable, Participants, Statut, Notes.
    """
    if not OPENPYXL_OK:
        return []

    try:
        wb = _get_workbook(path)
        ws, _ = _ensure_sheets(wb)

        meetings = []
        header = [str(c.value).strip() if c.value else '' for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v for v in row):
                continue
            m = {}
            for i, col in enumerate(MEETING_COLS):
                val = row[i] if i < len(row) else None
                m[col] = str(val).strip() if val is not None else ''
            # Format date if datetime object
            if isinstance(row[0], datetime.datetime):
                m['Date'] = row[0].strftime('%Y-%m-%d')
            meetings.append(m)

        return meetings
    except Exception as e:
        print(f"[manager_plan] Error loading meetings: {e}")
        return []


def save_meeting(path: str, meeting: dict) -> bool:
    """
    Append a new meeting row to the Excel file.
    meeting = {Date, Titre, Responsable, Participants, Statut, Notes}
    """
    if not OPENPYXL_OK:
        return False

    try:
        wb = _get_workbook(path)
        ws, _ = _ensure_sheets(wb)

        row = [
            meeting.get('Date', datetime.date.today().isoformat()),
            meeting.get('Titre', ''),
            meeting.get('Responsable', ''),
            meeting.get('Participants', ''),
            meeting.get('Statut', 'Planifiée'),
            meeting.get('Notes', ''),
        ]
        ws.append(row)
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        wb.save(path)
        return True
    except Exception as e:
        print(f"[manager_plan] Error saving meeting: {e}")
        return False


def update_meeting_status(path: str, date_str: str, title: str, new_status: str) -> bool:
    """Update the Statut of a meeting matched by Date + Titre."""
    if not OPENPYXL_OK:
        return False
    try:
        wb = _get_workbook(path)
        ws, _ = _ensure_sheets(wb)
        for row in ws.iter_rows(min_row=2):
            d_val = str(row[0].value).strip() if row[0].value else ''
            t_val = str(row[1].value).strip() if row[1].value else ''
            if d_val == date_str and t_val == title:
                row[4].value = new_status
                wb.save(path)
                return True
        return False
    except Exception as e:
        print(f"[manager_plan] Error updating meeting: {e}")
        return False


# ──────────────────────────────────────────────────────────────────────────────
# Actions
# ──────────────────────────────────────────────────────────────────────────────

def load_actions(path: str) -> list:
    """
    Load all action items from the Excel file.
    Returns list of dicts with keys: ID, Action, Responsable, Échéance, Statut, Notes, Réunion_Ref.
    """
    if not OPENPYXL_OK:
        return []

    try:
        wb = _get_workbook(path)
        _, ws = _ensure_sheets(wb)

        actions = []
        today = datetime.date.today()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v for v in row):
                continue
            a = {}
            for i, col in enumerate(ACTION_COLS):
                val = row[i] if i < len(row) else None
                a[col] = str(val).strip() if val is not None else ''

            # Handle date objects
            if isinstance(row[3], (datetime.datetime, datetime.date)):
                a['Échéance'] = row[3].strftime('%Y-%m-%d') if isinstance(row[3], datetime.datetime) else row[3].isoformat()

            # Compute overdue flag
            try:
                deadline = datetime.date.fromisoformat(a['Échéance'])
                a['overdue'] = (deadline < today) and a['Statut'] not in ('Terminée',)
            except Exception:
                a['overdue'] = False

            # Auto-generate ID if missing
            if not a.get('ID'):
                a['ID'] = str(uuid.uuid4())[:8].upper()

            actions.append(a)

        return actions
    except Exception as e:
        print(f"[manager_plan] Error loading actions: {e}")
        return []


def save_action(path: str, action: dict) -> bool:
    """
    Append a new action row. Auto-generates an ID if not provided.
    action = {Action, Responsable, Échéance, Statut, Notes, Réunion_Ref}
    """
    if not OPENPYXL_OK:
        return False

    try:
        wb = _get_workbook(path)
        _, ws = _ensure_sheets(wb)

        action_id = action.get('ID') or str(uuid.uuid4())[:8].upper()
        row = [
            action_id,
            action.get('Action', ''),
            action.get('Responsable', ''),
            action.get('Échéance', ''),
            action.get('Statut', 'Backlog'),
            action.get('Notes', ''),
            action.get('Réunion_Ref', ''),
        ]
        ws.append(row)
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        wb.save(path)
        return True
    except Exception as e:
        print(f"[manager_plan] Error saving action: {e}")
        return False


def update_action_status(path: str, action_id: str, new_status: str) -> bool:
    """Update the Statut column for the action with the given ID."""
    if not OPENPYXL_OK:
        return False
    try:
        wb = _get_workbook(path)
        _, ws = _ensure_sheets(wb)
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value).strip() == action_id:
                row[4].value = new_status
                wb.save(path)
                return True
        return False
    except Exception as e:
        print(f"[manager_plan] Error updating action: {e}")
        return False


def delete_action(path: str, action_id: str) -> bool:
    """Remove the row with the given action ID."""
    if not OPENPYXL_OK:
        return False
    try:
        wb = _get_workbook(path)
        _, ws = _ensure_sheets(wb)
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[0].value).strip() == action_id:
                ws.delete_rows(i)
                wb.save(path)
                return True
        return False
    except Exception as e:
        print(f"[manager_plan] Error deleting action: {e}")
        return False


# ──────────────────────────────────────────────────────────────────────────────
# KPI Summary
# ──────────────────────────────────────────────────────────────────────────────

def get_action_kpis(actions: list) -> dict:
    """Compute summary KPIs from the loaded actions list."""
    total    = len(actions)
    done     = sum(1 for a in actions if a.get('Statut') == 'Terminée')
    inprog   = sum(1 for a in actions if a.get('Statut') == 'En cours')
    overdue  = sum(1 for a in actions if a.get('overdue'))
    rate     = round(done / total * 100) if total else 0

    return {
        'total':   total,
        'done':    done,
        'inprog':  inprog,
        'overdue': overdue,
        'rate':    rate,
    }
