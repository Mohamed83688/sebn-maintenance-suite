"""
IMA — Excel Calendrier Reader & Machine Catalogue Export
Reads the SEBN-TN maintenance schedule Excel to extract unique machines.
Completely independent from PMA's DataEngine.
"""
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class CalendrierReader:
    """Reads SEBN-TN maintenance schedule Excel files and extracts machine data."""

    # Blacklisted sheet names (config / legend / meta sheets)
    SHEET_BLACKLIST = [
        "SOMMAIRE", "HELP", "LEGENDE", "SIGNATURES", "SIGNATURE",
        "USER", "USERS", "ACCOUNT", "LOGIN", "PARAM", "CONFIG", "BACKUP", "TEMPLATE", "FEUIL"
    ]

    def __init__(self):
        self.machines_df: pd.DataFrame | None = None
        self.source_path: str | None = None

    # ──────────────────────────────────────────────────────────────────
    #  CORE: Read Calendrier → Extract Unique Machines
    # ──────────────────────────────────────────────────────────────────
    def read_calendrier(self, file_path: str) -> pd.DataFrame:
        """
        Read the maintenance schedule Excel file.
        Returns a DataFrame with columns: Groupe, ID Machine, Nom Machine
        Each row is a unique machine found across all valid sheets.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Fichier introuvable : {file_path}")

        ext = os.path.splitext(file_path)[1].lower()

        # Choose engine
        try:
            engine = "calamine"
            pd.read_excel(file_path, engine=engine, nrows=1)
        except Exception:
            engine = "openpyxl" if ext in [".xlsx", ".xlsm"] else "xlrd"

        all_machines = []

        with pd.ExcelFile(file_path, engine=engine) as xl:
            sheets = [
                s for s in xl.sheet_names
                if not any(b in s.upper().strip() for b in self.SHEET_BLACKLIST)
                and s.upper().strip() != "TABLE" and not s.upper().strip().startswith("TABLE ")
            ]

            for sheet_name in sheets:
                try:
                    df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
                    if df.empty or len(df) < 2:
                        continue

                    # Find week start column and the header row index
                    week_start_col = None
                    header_row_idx = None
                    for i in range(min(20, len(df))):
                        row_vals = [str(x).upper().strip() for x in df.iloc[i]]
                        for j, val in enumerate(row_vals):
                            if val in ["S1", "S01", "W1", "W01", "KW 1", "KW1", "KW01", "KW 01", "KW.1", "WEEK 1"]:
                                week_start_col = j
                                header_row_idx = i
                                break
                            elif val.startswith("KW ") or val.startswith("S ") or val.startswith("W "):
                                suffix = val.split(" ")[-1]
                                if suffix.isdigit() and int(suffix) == 1:
                                    week_start_col = j
                                    header_row_idx = i
                                    break
                        if header_row_idx is not None:
                            break

                    if header_row_idx is None or week_start_col is None:
                        continue

                    # Identify exact columns
                    hdr_row = [str(df.iloc[header_row_idx, c] or '').strip() for c in range(week_start_col)]
                    col_machine = None
                    col_name = None
                    col_group = None

                    for c_i, h_val in enumerate(hdr_row):
                        h_up = h_val.upper()
                        if any(k in h_up for k in ['SEMAINE', 'ID MACHINE', 'MACHINE', 'EQUIPMENT', 'EQUIPEMENT']):
                            col_machine = c_i
                        elif h_up in ['N°', 'NO', 'DESIGNATION', 'NOM MACHINE']:
                            col_name = c_i
                        elif any(k in h_up for k in ['ZONE', 'GROUPE', 'GROUP']):
                            col_group = c_i

                    if col_machine is None:
                        col_machine = week_start_col - 1

                    # Scan rows below header
                    for i in range(header_row_idx + 1, len(df)):
                        row = df.iloc[i]
                        if len(row) <= week_start_col:
                            continue

                        raw_m = str(row[col_machine]).strip() if col_machine < len(row) and pd.notna(row[col_machine]) else ""
                        if not raw_m or raw_m.lower() in ["nan", "none", ""]:
                            continue

                        raw_m_up = raw_m.upper()
                        noise = [
                            'ZONE', 'N° CARTE', 'SEMAINE', 'TOTAL', 'ROLE', 'ADMIN', 'ADMINISTRATEUR',
                            'SIGNATURES', 'SIGNATURE', 'TECHNICIAN', 'TECHNICIEN', 'USER', 'USERS',
                            'ACCOUNT', 'LOGIN', 'MATRICULE', 'DATE', 'SHIFT', 'EQUIPE', 'PAGE',
                            'RESP', 'REV', 'ANNEXE', 'SOMMAIRE', 'VALIDÉ', 'VALIDE', 'VISA'
                        ]
                        if raw_m_up in noise or raw_m_up.startswith('PPE-VA') or raw_m_up.startswith('ANNEXE'):
                            continue

                        # Check if has schedule markers
                        schedule_area = row[week_start_col:min(week_start_col + 54, len(row))]
                        has_markers = any(
                            any(x in str(v).strip().upper() for x in ['M', 'H', 'DONE', 'X', 'OK'])
                            for v in schedule_area if pd.notna(v)
                        )
                        if not has_markers:
                            continue

                        zone_val = str(row[col_group]).strip() if col_group is not None and col_group < len(row) and pd.notna(row[col_group]) else ""
                        group_name = zone_val if zone_val and zone_val.lower() != 'nan' else sheet_name.strip()

                        m_name_val = str(row[col_name]).strip() if col_name is not None and col_name < len(row) and pd.notna(row[col_name]) else ""
                        machine_name = m_name_val if m_name_val and m_name_val.lower() != 'nan' else raw_m

                        all_machines.append({
                            "Groupe": group_name,
                            "ID Machine": raw_m,
                            "Nom Machine": machine_name
                        })
                except Exception:
                    continue

        if not all_machines:
            self.machines_df = pd.DataFrame(columns=["Groupe", "ID Machine"])
            return self.machines_df

        raw_df = pd.DataFrame(all_machines)

        # Deduplicate
        self.machines_df = (
            raw_df.drop_duplicates(subset=["Groupe", "ID Machine"])
            .sort_values(["Groupe", "ID Machine"])
            .reset_index(drop=True)
        )
        self.source_path = file_path
        return self.machines_df

    def read_asp_codes(self, file_path: str) -> list[dict]:
        """
        Reads SAP/ASP codes and descriptions from an Excel file.
        Inclusive logic to ensure no codes are missed.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Fichier introuvable : {file_path}")

        all_codes = []
        try:
            # Using dtype=str ensures Excel numbers aren't converted to floats with .0
            with pd.ExcelFile(file_path) as xl:
                for sheet_name in xl.sheet_names:
                    df = pd.read_excel(xl, sheet_name=sheet_name, header=None, dtype=str)
                    if df.empty: continue

                    # Identify Columns
                    code_col = 0
                    desc_col = 2 # Default: Column A=Code, C=Description
                    header_row_idx = -1

                    # Look for headers in the first 30 rows
                    for i in range(min(30, len(df))):
                        row_vals = [str(x).upper().strip() for x in df.iloc[i]]
                        # The user's image shows "SAP TN30" and "Désigantion"
                        if any("SAP" in v or "CODE" in v for v in row_vals):
                            if any("DÉSIG" in v or "DESIGN" in v or "NOM" in v for v in row_vals):
                                for j, val in enumerate(row_vals):
                                    if "SAP" in val or "CODE" in val: code_col = j
                                    if "DÉSIG" in val or "DESIGN" in val or "NOM" in val: desc_col = j
                                header_row_idx = i
                                break
                    
                    # If we found a header, start reading immediately after it.
                    # Otherwise, use the user's "ligne 4" rule (index 3).
                    start_idx = header_row_idx + 1 if header_row_idx != -1 else 3
                    
                    # Deduplication set for this sheet
                    seen_in_sheet = set()

                    for i in range(start_idx, len(df)):
                        row = df.iloc[i]
                        if len(row) <= max(code_col, desc_col): continue
                        
                        code = str(row[code_col]).strip()
                        # Skip clearly non-code rows
                        if not code or code.lower() in ["nan", "none", "null", "total", "somme"]:
                            continue
                        
                        # Fix: Don't skip if it just CONTAINS SAP TN, only if it IS the header
                        if "SAP TN" in code.upper() and len(code) < 15:
                             # Check if it's the header row again (redundancy check)
                             if "DÉSIG" in str(row[desc_col]).upper(): continue

                        desc = str(row[desc_col]).strip() if desc_col < len(row) else ""
                        if desc.lower() == "nan": desc = ""
                        
                        if code not in seen_in_sheet:
                            all_codes.append({"code": code, "description": desc})
                            seen_in_sheet.add(code)
                        
        except Exception:
            pass
            
        return all_codes

    # ──────────────────────────────────────────────────────────────────
    #  PRIVATE: Machine-ID Detection Heuristic
    # ──────────────────────────────────────────────────────────────────
    def _identify_machine(self, row, week_start_col: int, sheet_name: str) -> dict | None:
        """Find machine ID and name in the columns before week columns."""
        # SCAN RIGHT TO LEFT: Pick the most specific name closest to the markers
        search_range = list(reversed(range(0, week_start_col)))
        
        found_id = None
        found_name = None

        for col_idx in search_range:
            val = str(row[col_idx]).strip()
            # Skip empty or group-title matches
            if not val or val.lower() == "nan" or val.upper() == sheet_name.upper():
                continue

            # Len >= 1 to capture everything including short numeric IDs
            if len(val) >= 1 and val.count(" ") <= 4:
                blacklist = [
                    "ZONE", "ANNÉE", "CARTE", "RESP", "REV", "TYPE", "MODÈLE", 
                    "DATE", "SÉRIE", "PLAN", "RÉGION", "SECTEUR", "TOTAL", 
                    "NB", "MOIS", "HEURE", "CUMUL", "SEMAINE", "JOUR",
                    "DOMAINE", "SECTION", "SOMAIN"
                ]
                if not any(b in val.upper() for b in blacklist):
                    if not found_id:
                        found_id = val
                    elif not found_name:
                        found_name = val
                    else:
                        break

        if found_id:
            return {"id": found_id, "name": found_name or found_id}
        return None

    # ──────────────────────────────────────────────────────────────────
    #  EXPORT: Save Machine Catalogue to Excel
    # ──────────────────────────────────────────────────────────────────
    def export_catalogue(self, output_path: str,
                         machines_df: pd.DataFrame = None) -> str:
        """
        Export the machine catalogue as a professional Excel file.
        Creates: SOMMAIRE + one sheet per group.
        Returns the output path.
        """
        df = machines_df if machines_df is not None else self.machines_df
        if df is None or df.empty:
            raise ValueError("Aucune donnée machine à exporter")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Style constants
        HDR_FILL  = PatternFill("solid", fgColor="0D6EFD")
        ALT_FILL  = PatternFill("solid", fgColor="EBF3FF")
        WHT_FILL  = PatternFill("solid", fgColor="FFFFFF")
        HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        DATA_FONT = Font(name="Calibri", size=11)
        CENTER    = Alignment(horizontal="center", vertical="center")
        LEFT      = Alignment(horizontal="left", vertical="center")
        thin      = Side(style="thin", color="CCCCCC")
        BORDER    = Border(bottom=thin)

        def write_sheet(ws, grp_df, title):
            ws.title = title[:31]
            ws.sheet_view.showGridLines = False
            ws.row_dimensions[1].height = 30

            headers = ["GROUPE", "ID MACHINE"]
            widths  = [25, 30]

            for ci, (h, w) in enumerate(zip(headers, widths), 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = CENTER
                ws.column_dimensions[get_column_letter(ci)].width = w

            for ri, (_, row) in enumerate(grp_df.iterrows(), 2):
                ws.row_dimensions[ri].height = 22
                fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
                vals = [
                    row.get("Groupe", ""),
                    row.get("ID Machine", ""),
                ]
                for ci, val in enumerate(vals, 1):
                    cell = ws.cell(row=ri, column=ci, value=str(val) if val else "")
                    cell.fill = fill
                    cell.font = DATA_FONT
                    cell.border = BORDER
                    cell.alignment = CENTER

        # SOMMAIRE
        ws_sum = wb.create_sheet("SOMMAIRE")
        write_sheet(ws_sum, df, "SOMMAIRE")

        # One sheet per group
        for grp in sorted(df["Groupe"].unique()):
            grp_df = df[df["Groupe"] == grp]
            ws_grp = wb.create_sheet()
            write_sheet(ws_grp, grp_df, str(grp))

        wb.save(output_path)
        return output_path
